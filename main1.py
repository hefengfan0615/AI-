import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import pandas as pd
import re
from datetime import datetime
import warnings
import locale
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

# ---------- 1. 抑制 OpenPyXL 警告 ----------
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# ---------- 2. 修正 Tkinter 中文报错 ----------
try:
    locale.setlocale(locale.LC_CTYPE, 'Chinese')
except:
    try:
        locale.setlocale(locale.LC_CTYPE, 'zh_CN.UTF-8')
    except:
        pass

# ==================== 内置门店信息映射表 ====================
# 根据您提供的数据构建（Store Nbr → 省份, 大区, 门店名称, 城市）
STORE_INFO = {
    1009: ("吉林", "1区", "沃尔玛吉林长春前进广场店(总店)", "吉林"),
    2012: ("吉林", "1区", "沃尔玛吉林长春临河街店", "吉林"),
    2440: ("吉林", "1区", "沃尔玛吉林长春博学路店", "吉林"),
    3425: ("吉林", "1区", "沃尔玛吉林长春重庆路店", "吉林"),
    808: ("辽宁", "1区", "沃尔玛辽宁大连安盛店", "大连"),
    811: ("辽宁", "1区", "沃尔玛辽宁大连学苑店", "大连"),
    812: ("辽宁", "1区", "沃尔玛辽宁大连旅顺店", "大连"),
    1008: ("辽宁", "1区", "沃尔玛辽宁大连西安路店", "大连"),
    2067: ("辽宁", "1区", "沃尔玛辽宁沈阳长青", "沈阳"),
    2141: ("辽宁", "1区", "沃尔玛辽宁大连周水前店", "大连"),
    2397: ("辽宁", "1区", "沃尔玛辽宁大连金州店", "大连"),
    3407: ("辽宁", "1区", "沃尔玛辽宁大连奥林匹克店", "大连"),
    3414: ("辽宁", "1区", "沃尔玛辽宁大连华南店", "大连"),
    5780: ("辽宁", "1区", "沃尔玛辽宁沈阳中华路店", "沈阳"),
    2470: ("内蒙古", "1区", "沃尔玛内蒙古赤峰赤峰店", "赤峰"),
    2371: ("山东", "1区", "沃尔玛山东济宁济宁店", "济宁"),
    802: ("北京", "2区", "沃尔玛北京世纪城店", "北京"),
    804: ("北京", "2区", "沃尔玛北京五棵松店", "北京"),
    1034: ("北京", "2区", "沃尔玛北京万达广场店", "北京"),
    2028: ("北京", "2区", "沃尔玛北京昌平店", "北京"),
    2068: ("北京", "2区", "沃尔玛北京延庆县延庆店", "北京"),
    2784: ("北京", "2区", "BEIJING QINGHE BRANCH SC", "北京"),
    3423: ("北京", "2区", "沃尔玛北京宣武门店", "北京"),
    1077: ("河北", "2区", "沃尔玛购物广场河北廊坊朝阳分店", "廊坊"),
    2116: ("河北", "2区", "沃尔玛河北保定朝阳大街店", "保定"),
    2168: ("河北", "2区", "沃尔玛购物广场河北邯郸东柳大街分店", "邯郸"),
    2369: ("河北", "2区", "沃尔玛河北邢台中兴东大街店", "邢台"),
    2414: ("北京", "2区", "沃尔玛百货有限公司北京(燕郊店)", "北京"),
    2022: ("河南", "2区", "沃尔玛购物广场河南安阳人民大道分店", "安阳"),
    2154: ("河南", "2区", "沃尔玛河南信阳人民店", "信阳"),
    2379: ("河南", "2区", "沃尔玛购物广场河南郑州中原西路分店", "郑州"),
    2127: ("内蒙古", "2区", "沃尔玛购物广场内蒙包头文化路分店", "包头"),
    1010: ("山西", "2区", "沃尔玛山西太原长风街店", "太原"),
    1079: ("山西", "2区", "沃尔玛购物广场山西大同永泰分店", "大同"),
    2005: ("山西", "2区", "沃尔玛购物广场山西阳泉桃北东路分店", "阳泉"),
    2080: ("山西", "2区", "沃尔玛山西运城运城店", "运城"),
    2381: ("山西", "2区", "沃尔玛山西吕梁孝义市崇文店", "吕梁"),
    2533: ("山西", "2区", "沃尔玛山西晋中安宁街分店", "晋中"),
    1026: ("陕西", "2区", "沃尔玛购物广场陕西西安金花南路分店", "西安"),
    2044: ("陕西", "2区", "沃尔玛购物广场陕西西安雁塔路店", "西安"),
    2712: ("天津", "2区", "沃尔玛天津武清店", "天津"),
    3421: ("天津", "2区", "沃尔玛天津河东店", "天津"),
    5019: ("JD", "JD店", "京东到家", "JD"),
    2038: ("安徽", "3区", "沃尔玛安徽六安梅山路店", "六安"),
    2172: ("安徽", "3区", "沃尔玛安徽亳州魏武广场店", "亳州"),
    2204: ("安徽", "3区", "沃尔玛安徽合肥港澳广场店", "合肥"),
    309: ("江苏", "3区", "沃尔玛江苏苏州南门店", "苏州"),
    311: ("江苏", "3区", "沃尔玛江苏南京广鼎店", "南京"),
    1040: ("江苏", "3区", "沃尔玛江苏扬州嘉信茂店", "扬州"),
    1055: ("江苏", "3区", "沃尔玛江苏苏州太仓市", "苏州"),
    2057: ("江苏", "3区", "沃尔玛江苏无锡金城店", "无锡"),
    2093: ("江苏", "3区", "沃尔玛江苏南京浦口店", "南京"),
    2199: ("江苏", "3区", "沃尔玛江苏南京虹悦城店", "南京"),
    2430: ("江苏", "3区", "沃尔玛江苏常州万达店", "常州"),
    5020: ("JD", "JD店", "京东到家", "JD"),
    5823: ("江苏", "3区", "沃尔玛江苏南京新街口分店", "南京"),
    302: ("上海", "3区", "沃尔玛上海田林店", "上海"),
    310: ("上海", "3区", "沃尔玛上海凌云店", "上海"),
    1014: ("上海", "3区", "沃尔玛上海", "上海"),
    1080: ("上海", "3区", "沃尔玛上海松江店", "上海"),
    2056: ("上海", "3区", "沃尔玛上海高行店", "上海"),
    2076: ("上海", "3区", "沃尔玛上海桃浦店", "上海"),
    2165: ("上海", "3区", "沃尔玛上海顾村店", "上海"),
    2184: ("上海", "3区", "沃尔玛上海水产店", "上海"),
    2475: ("上海", "3区", "沃尔玛上海机场店", "上海"),
    703: ("浙江", "3区", "沃尔玛浙江宁波联丰店", "宁波"),
    710: ("浙江", "3区", "沃尔玛浙江嘉兴华庭店", "嘉兴"),
    711: ("浙江", "3区", "沃尔玛浙江温州鸿英店", "温州"),
    714: ("浙江", "3区", "沃尔玛浙江温州欧洲城店", "温州"),
    1039: ("浙江", "3区", "沃尔玛浙江金华福华广场店", "金华"),
    1074: ("浙江", "3区", "沃尔玛浙江宁波万达店", "宁波"),
    1078: ("浙江", "3区", "沃尔玛浙江杭州临平店", "杭州"),
    2037: ("浙江", "3区", "沃尔玛浙江金华浦江县浦江店", "金华"),
    2043: ("浙江", "3区", "沃尔玛浙江湖州德清县武康店", "湖州"),
    2065: ("浙江", "3区", "沃尔玛浙江杭州三墩店", "杭州"),
    2182: ("浙江", "3区", "沃尔玛浙江金华武义县俞源店", "金华"),
    2399: ("浙江", "3区", "沃尔玛浙江杭州临安市钱王街店", "杭州"),
    2458: ("浙江", "3区", "沃尔玛浙江温州乐清市伯乐东路店", "温州"),
    2487: ("浙江", "3区", "沃尔玛浙江金华兰溪市兰溪店", "金华"),
    2569: ("浙江", "3区", "沃尔玛浙江绍兴新昌县海洋城店", "绍兴"),
    2598: ("浙江", "3区", "沃尔玛浙江温州平阳县鳌江店", "温州"),
    2599: ("浙江", "3区", "沃尔玛浙江温州苍南县苍南店", "温州"),
    1003: ("湖北", "4区", "沃尔玛湖北武汉中山大道店", "武汉"),
    1013: ("湖北", "4区", "沃尔玛湖北武汉徐东大街店", "武汉"),
    1051: ("湖北", "4区", "沃尔玛湖北荆州北京中路分店", "荆州"),
    1065: ("湖北", "4区", "沃尔玛湖北襄阳长虹路分店", "襄阳"),
    2041: ("湖北", "4区", "沃尔玛湖北武汉西汇店", "武汉"),
    2073: ("湖北", "4区", "沃尔玛湖北武汉王家湾店", "武汉"),
    2101: ("湖北", "4区", "沃尔玛湖北武汉南湖店", "武汉"),
    2130: ("湖北", "4区", "沃尔玛湖北宜昌夷陵广场分店", "宜昌"),
    2142: ("湖北", "4区", "沃尔玛湖北武汉钟家村店", "武汉"),
    2156: ("湖北", "4区", "沃尔玛湖北武汉光谷店", "武汉"),
    2191: ("湖北", "4区", "沃尔玛湖北武汉黄陂广场店", "武汉"),
    2201: ("湖北", "4区", "沃尔玛湖北宜昌九码头分店", "宜昌"),
    2202: ("湖北", "4区", "沃尔玛湖北武汉菱角湖万达店", "武汉"),
    2376: ("湖北", "4区", "沃尔玛湖北武汉奥山世纪城店", "武汉"),
    2401: ("湖北", "4区", "沃尔玛湖北黄石黄石店", "黄石"),
    2415: ("湖北", "4区", "沃尔玛湖北咸宁金城店", "咸宁"),
    2420: ("湖北", "4区", "沃尔玛湖北武汉经开店", "武汉"),
    2426: ("湖北", "4区", "沃尔玛湖北黄冈麻城市麻城广场店", "黄冈"),
    2461: ("湖北", "4区", "沃尔玛湖北黄冈红安县红安店", "黄冈"),
    2485: ("湖北", "4区", "沃尔玛湖北天门市天门店", "天门"),
    2579: ("湖北", "4区", "沃尔玛湖北荆州洪湖市洪湖店", "荆州"),
    2580: ("湖北", "4区", "沃尔玛湖北孝感汉川市西正街店", "孝感"),
    2581: ("湖北", "4区", "沃尔玛湖北武汉吴家山店", "武汉"),
    2590: ("湖北", "4区", "沃尔玛湖北孝感应城市应城分店", "孝感"),
    2594: ("湖北", "4区", "沃尔玛湖北仙桃市仙桃店", "仙桃"),
    2702: ("湖北", "4区", "沃尔玛湖北武汉杨家湾店", "武汉"),
    2704: ("湖北", "4区", "沃尔玛湖北武汉江夏店", "武汉"),
    2746: ("湖北", "4区", "沃尔玛湖北武汉马鹦路店", "武汉"),
    2748: ("湖北", "4区", "沃尔玛湖北宜昌环球港店", "宜昌"),
    2752: ("湖北", "4区", "沃尔玛湖北武汉印象城店", "武汉"),
    2758: ("湖北", "4区", "沃尔玛购物广场湖北孝感安陆市太白大道分店", "孝感"),
    5021: ("JD", "JD店", "京东到家", "JD"),
    1006: ("湖南", "4区", "沃尔玛湖南长沙雨花亭", "长沙"),
    1036: ("湖南", "4区", "沃尔玛湖南岳阳岳阳店", "岳阳"),
    1066: ("湖南", "4区", "沃尔玛湖南娄底春园店", "娄底"),
    1087: ("湖南", "4区", "沃尔玛湖南娄底大汉店", "娄底"),
    2015: ("湖南", "5区", "沃尔玛湖南郴州郴州店", "郴州"),   # 原4区，根据数据修正
    2025: ("湖南", "4区", "沃尔玛湖南益阳益阳店", "益阳"),
    2149: ("湖南", "4区", "沃尔玛湖南邵阳邵阳店", "邵阳"),
    2410: ("湖南", "4区", "沃尔玛湖南株洲攸县攸县店", "株洲"),
    2444: ("湖南", "4区", "沃尔玛湖南株洲神龙城店", "株洲"),
    2445: ("湖南", "4区", "沃尔玛湖南娄底新化县新化店", "娄底"),
    2562: ("湖南", "4区", "沃尔玛湖南长沙梅溪湖店", "长沙"),
    2710: ("湖南", "4区", "沃尔玛湖南长沙开福店", "长沙"),
    2727: ("湖南", "4区", "沃尔玛湖南岳阳青年东路店", "岳阳"),
    2738: ("湖南", "4区", "沃尔玛湖南长沙观沙路店", "长沙"),
    2753: ("湖南", "4区", "沃尔玛湖南长沙学士店", "长沙"),
    5746: ("湖南", "4区", "沃尔玛湖南长沙黄兴店", "长沙"),
    612: ("江西", "4区", "沃尔玛江西南昌三店西路店", "南昌"),
    1053: ("江西", "4区", "沃尔玛江西景德镇广场南路店", "景德镇"),
    1099: ("江西", "4区", "沃尔玛江西九江庐山南路店", "九江"),
    2010: ("江西", "4区", "沃尔玛江西鹰潭站江店", "鹰潭"),
    2479: ("江西", "4区", "沃尔玛江西吉安吉安店", "吉安"),
    2538: ("江西", "4区", "沃尔玛江西南昌新建店", "南昌"),
    2585: ("江西", "4区", "沃尔玛购物广场南昌象湖分店", "南昌"),
    2713: ("江西", "4区", "江西赣州八四一大道分店", "赣州"),
    2714: ("江西", "4区", "沃尔玛购物广场新赣州大道分店", "赣州"),
    5782: ("江西", "4区", "沃尔玛江西南昌八一店(总店)", "南昌"),
    602: ("福建", "5区", "沃尔玛福建福州福清市融侨", "福州"),
    610: ("福建", "5区", "沃尔玛福建福州富贵店", "福州"),
    611: ("福建", "5区", "沃尔玛福建厦门聚祥店", "厦门"),
    1019: ("福建", "5区", "沃尔玛福建福州鼓山店", "福州"),
    1025: ("福建", "5区", "沃尔玛福建漳州丹霞店", "漳州"),
    1035: ("福建", "5区", "沃尔玛福建泉州晋江市晋江店", "泉州"),
    2020: ("福建", "5区", "沃尔玛福建厦门加州店", "厦门"),
    2050: ("福建", "5区", "沃尔玛福建莆田文献店", "莆田"),
    2058: ("福建", "5区", "沃尔玛福建厦门瑞景店", "厦门"),
    2059: ("福建", "5区", "沃尔玛福建福州则徐店", "福州"),
    2071: ("福建", "5区", "沃尔玛福建福州长乐店", "福州"),
    2098: ("福建", "5区", "沃尔玛福建宁德宁德店", "宁德"),
    2351: ("福建", "5区", "沃尔玛福建厦门海沧店", "厦门"),
    2380: ("福建", "5区", "沃尔玛福建厦门湖里万达店", "厦门"),
    2403: ("福建", "5区", "沃尔玛福建福州仓山万达店", "福州"),
    2412: ("福建", "5区", "沃尔玛福建宁德福鼎市福鼎店", "宁德"),
    2449: ("福建", "5区", "沃尔玛福建福州秀峰店", "福州"),
    2480: ("福建", "5区", "沃尔玛福建宁德霞浦县霞浦店", "宁德"),
    2543: ("福建", "5区", "沃尔玛福建莆田涵江店", "莆田"),
    2564: ("福建", "5区", "沃尔玛购物广场福建福州连江县连江店", "福州"),
    2729: ("福建", "5区", "沃尔玛福建莆田名邦豪苑店", "莆田"),
    2740: ("福建", "5区", "沃尔玛福建泉州笋江店", "泉州"),
    3410: ("福建", "5区", "沃尔玛福建福州利嘉店", "福州"),
    3413: ("福建", "5区", "沃尔玛福建福州长城店", "福州"),
    3417: ("福建", "5区", "沃尔玛福建厦门SM店", "厦门"),
    3420: ("福建", "5区", "沃尔玛福建厦门世贸店", "厦门"),
    101: ("广东", "5区", "沃尔玛广东广州天河店", "广州"),
    105: ("广东", "5区", "沃尔玛广东广州淘金店", "广州"),
    107: ("广东", "5区", "沃尔玛广东深圳红岭店", "深圳"),
    111: ("广东", "5区", "沃尔玛广东广州黄埔店", "广州"),
    114: ("广东", "5区", "沃尔玛广东湛江赤坎店", "外埠"),
    117: ("广东", "5区", "沃尔玛广东广州新港店", "广州"),
    118: ("广东", "5区", "沃尔玛广东东莞虎门店", "东莞"),
    119: ("广东", "5区", "沃尔玛广东深圳翠竹店", "深圳"),
    120: ("广东", "5区", "沃尔玛广东广州南洲店", "广州"),
    121: ("广东", "5区", "沃尔玛广东深圳大益店", "深圳"),
    122: ("广东", "5区", "沃尔玛广东佛山怡东店", "佛山"),
    124: ("广东", "5区", "沃尔玛广东广州芳村店", "广州"),
    126: ("广东", "5区", "沃尔玛广东东莞世博店", "东莞"),
    127: ("广东", "5区", "沃尔玛广东佛山季华2店", "佛山"),
    128: ("广东", "5区", "沃尔玛广东东莞长安店", "东莞"),
    1022: ("广东", "5区", "沃尔玛广东佛山桂城店", "佛山"),
    1029: ("广东", "5区", "沃尔玛广东佛山季华1店", "佛山"),
    1030: ("广东", "5区", "沃尔玛广东东莞鸿福店", "东莞"),
    1037: ("广东", "5区", "沃尔玛广东湛江赤坎店", "外埠"),
    1038: ("广东", "5区", "沃尔玛广东惠州惠阳店", "惠州"),
    1043: ("广东", "5区", "沃尔玛深圳龙岗（平岗中学）店", "深圳"),
    1047: ("广东", "5区", "沃尔玛广东东莞长安广场店", "东莞"),
    1048: ("广东", "5区", "沃尔玛广东茂名茂名店", "外埠"),
    1056: ("广东", "5区", "沃尔玛广东深圳华强北店", "深圳"),
    1059: ("广东", "5区", "沃尔玛广东深圳香梅店", "深圳"),
    1070: ("广东", "5区", "沃尔玛广东惠州惠州店", "惠州"),
    1082: ("广东", "5区", "沃尔玛广东韶关新华店", "外埠"),
    1094: ("广东", "5区", "沃尔玛广东深圳罗田店", "深圳"),
    1095: ("广东", "5区", "沃尔玛广东湛江霞山店", "外埠"),
    2016: ("广东", "5区", "沃尔玛广东肇庆端州店", "外埠"),
    2026: ("广东", "5区", "沃尔玛广东佛山顺德店", "佛山"),
    2027: ("广东", "5区", "沃尔玛广东深圳金港店", "深圳"),
    2066: ("广东", "5区", "沃尔玛广东珠海珠海店", "中珠"),
    2072: ("广东", "5区", "沃尔玛广东东莞凤岗中心店", "东莞"),
    2075: ("广东", "5区", "沃尔玛广东东莞黄江店", "东莞"),
    2082: ("广东", "5区", "沃尔玛广东中山张家边店", "中珠"),
    2153: ("广东", "5区", "沃尔玛广东东莞宏图店", "东莞"),
    2160: ("广东", "5区", "沃尔玛广东清远先锋东路店", "外埠"),
    2163: ("广东", "5区", "沃尔玛广东清远英德市英德店", "外埠"),
    2171: ("广东", "5区", "沃尔玛广东东莞塘厦店", "东莞"),
    2179: ("广东", "5区", "沃尔玛广东深圳沙井店", "深圳"),
    2181: ("广东", "5区", "沃尔玛广东湛江吴川店", "外埠"),
    2186: ("广东", "5区", "沃尔玛深圳龙华店", "深圳"),
    2209: ("广东", "5区", "沃尔玛广东广州白云万达店", "广州"),
    2218: ("广东", "5区", "沃尔玛广东佛山高明店", "佛山"),
    2220: ("广东", "5区", "沃尔玛广东珠海前山店", "中珠"),
    2232: ("广东", "5区", "沃尔玛广东惠州江北店", "惠州"),
    2340: ("广东", "5区", "沃尔玛广东惠州鹅岭店", "惠州"),
    2363: ("广东", "5区", "沃尔玛广东肇庆肇庆店", "外埠"),
    2377: ("广东", "5区", "沃尔玛广东湛江遂溪店", "外埠"),
    2406: ("广东", "5区", "沃尔玛广东汕头星湖商业城店", "潮汕"),
    2447: ("广东", "5区", "沃尔玛广东深圳横岗六约店", "深圳"),
    2463: ("广东", "5区", "沃尔玛广东惠州惠阳澳头店", "惠州"),
    2486: ("广东", "5区", "沃尔玛广东东莞大岭山玉屏路店", "东莞"),
    2489: ("广东", "5区", "沃尔玛广东深圳龙岗（峦山谷）店", "深圳"),
    2493: ("广东", "5区", "沃尔玛广东深圳龙岗建设店", "深圳"),
    2497: ("广东", "5区", "沃尔玛广东惠州仲恺大道分店", "惠州"),
    2502: ("广东", "5区", "沃尔玛广东深圳布吉锦龙路分店", "深圳"),
    2506: ("广东", "5区", "沃尔玛广东深圳东方大道分店", "深圳"),
    2510: ("广东", "5区", "沃尔玛广东河源中山大道分店", "外埠"),
    2530: ("广东", "5区", "沃尔玛广东东莞石龙店", "东莞"),
    2539: ("广东", "5区", "沃尔玛广东阳江西平北路分店", "外埠"),
    2547: ("广东", "5区", "沃尔玛广东阳江东风路分店", "外埠"),
    2555: ("广东", "5区", "沃尔玛购物广场广东深圳宝安新安路分店", "深圳"),
    2556: ("广东", "5区", "沃尔玛广东湛江廉江市百信广场分店", "外埠"),
    2582: ("广东", "5区", "沃尔玛广东河源越王大道分店", "外埠"),
    2586: ("广东", "5区", "沃尔玛广东深圳荣德国际店", "深圳"),
    2708: ("广东", "5区", "沃尔玛购物广场广东汕头东厦路分店", "潮汕"),
    2717: ("广东", "5区", "沃尔玛广东佛山环宇城分店", "佛山"),
    2723: ("广东", "5区", "沃尔玛广东揭阳普宁市广达北路分店", "潮汕"),
    2730: ("广东", "5区", "沃尔玛购物广场汕头嵩山路分店", "潮汕"),
    2733: ("广东", "5区", "沃尔玛广东深圳龙岗坪地分店", "深圳"),
    2735: ("广东", "5区", "沃尔玛广东深圳光明长升路分店", "深圳"),
    2743: ("广东", "5区", "沃尔玛广东惠州大亚湾龙光城分店", "惠州"),
    2744: ("广东", "5区", "沃尔玛广东广州钟村分店", "广州"),
    2754: ("广东", "5区", "沃尔玛购物广场广东汕尾香洲西路分店", "潮汕"),
    2756: ("广东", "5区", "沃尔玛购物广场广东东莞太沙路分店", "东莞"),
    2762: ("广东", "5区", "沃尔玛广东东莞香市路分店", "东莞"),
    2768: ("广东", "5区", "沃尔玛广东佛山大沥分店", "佛山"),
    3400: ("广东", "5区", "沃尔玛广东东莞东湖店", "东莞"),
    3402: ("广东", "5区", "沃尔玛广东深圳凤凰店", "深圳"),
    3403: ("广东", "5区", "沃尔玛广东深圳福星店", "深圳"),
    3404: ("广东", "5区", "沃尔玛广东深圳蛇口店", "深圳"),
    3409: ("广东", "5区", "沃尔玛广东深圳华侨城店", "深圳"),
    3432: ("广东", "5区", "沃尔玛广东深圳前进店", "深圳"),
    5018: ("JD", "JD店", "京东到家", "JD"),
    5016: ("JD", "JD店", "京东到家", "JD"),
    5824: ("广东", "5区", "沃尔玛广东深圳布吉店", "深圳"),
    1002: ("广西", "5区", "沃尔玛广西南宁朝阳店", "南宁"),
    1088: ("广西", "5区", "沃尔玛广西南宁航洋店", "南宁"),
    2002: ("广西", "5区", "沃尔玛广西桂林中山店", "桂林"),
    2119: ("广西", "5区", "沃尔玛广西北海北海店", "北海"),
    2233: ("广西", "5区", "沃尔玛广西南宁正恒店", "南宁"),
    2352: ("广西", "5区", "沃尔玛广西钦州钦州店", "钦州"),
    2427: ("广西", "5区", "沃尔玛来宾桂中大道店", "来宾"),
    2443: ("广西", "5区", "沃尔玛广西梧州梧州店", "梧州"),
    2494: ("广西", "5区", "沃尔玛广西南宁永和店", "南宁"),
    2505: ("广西", "5区", "沃尔玛广西南宁望州店", "南宁"),
    2567: ("广西", "5区", "沃尔玛购物广场广西防城港防邕路分店", "防城港"),
    2718: ("广西", "5区", "沃尔玛广西柳州柳州店", "柳州"),
    2719: ("广西", "5区", "沃尔玛广西柳州华润路分店", "柳州"),
    2745: ("广西", "5区", "沃尔玛购物广场广西桂林红岭分店", "桂林"),
    2751: ("广西", "5区", "沃尔玛购物广场广西南宁明秀东路分店", "南宁"),
    1000: ("贵州", "6区", "贵阳沙冲路分店", "贵阳"),
    1001: ("贵州", "6区", "贵阳人民广场分店", "贵阳"),
    1052: ("贵州", "6区", "六盘水钟山中路分店", "六盘水"),
    2032: ("贵州", "6区", "遵义纪念广场分店", "遵义"),
    2102: ("贵州", "6区", "沃尔玛贵州贵阳大营坡商业广场店", "贵阳"),
    2143: ("贵州", "6区", "沃尔玛贵州贵阳黄河路店", "贵阳"),
    2225: ("贵州", "6区", "沃尔玛购物广场贵州贵阳林城西路分店", "贵阳"),
    2515: ("贵州", "6区", "沃尔玛贵州毕节文博路分店", "毕节"),
    2559: ("贵州", "6区", "沃尔玛购物广场都匀斗篷山路分店", "都匀"),
    2765: ("贵州", "6区", "沃尔玛购物广场花溪甲秀南路分店", "贵阳"),
    201: ("四川", "6区", "沃尔玛四川成都亚太店", "成都"),
    205: ("四川", "6区", "沃尔玛四川绵阳兴达店", "绵阳"),
    206: ("四川", "6区", "沃尔玛四川成都府河店", "成都"),
    207: ("四川", "6区", "沃尔玛四川成都武侯店", "成都"),
    211: ("四川", "6区", "沃尔玛四川南充五星店（211店）", "南充"),
    213: ("四川", "6区", "好又多四川成都西城店", "成都"),
    215: ("四川", "6区", "沃尔玛四川成都双流县双流店", "成都"),
    1015: ("四川", "6区", "沃尔玛四川成都九里堤店", "成都"),
    1023: ("四川", "6区", "沃尔玛四川绵阳一店", "绵阳"),
    1044: ("四川", "6区", "沃尔玛四川成都SM店", "成都"),
    1045: ("四川", "6区", "沃尔玛四川宜宾酒都路店", "宜宾"),
    2018: ("四川", "6区", "沃尔玛四川达州朝阳东路店", "达州"),
    2029: ("四川", "6区", "沃尔玛四川德阳天山北路店", "德阳"),
    2031: ("四川", "6区", "沃尔玛四川遂宁嘉禾店", "遂宁"),
    2087: ("四川", "6区", "沃尔玛四川乐山土桥街店", "乐山"),
    2100: ("四川", "6区", "沃尔玛四川凉山西昌市航天大道店", "凉山"),
    2131: ("四川", "6区", "沃尔玛四川成都彭州市金彭东路店", "成都"),
    2135: ("四川", "6区", "沃尔玛四川攀枝花机场路店", "攀枝花"),
    2150: ("四川", "6区", "沃尔玛四川成都中海国际成都店", "成都"),
    2157: ("四川", "6区", "沃尔玛四川绵阳江油市白路店", "绵阳"),
    2216: ("四川", "6区", "沃尔玛四川自贡汇兴路店", "自贡"),
    2336: ("四川", "6区", "沃尔玛四川南充阆中市人民广场店", "南充"),
    2353: ("四川", "6区", "沃尔玛四川巴中巴中店", "巴中"),
    2382: ("四川", "6区", "沃尔玛四川成都华阳店", "成都"),
    2408: ("四川", "6区", "沃尔玛四川绵阳1958店", "绵阳"),
    2239: ("四川", "6区", "沃尔玛四川绵阳绵阳市江油李白大道分店", "绵阳"),
    2504: ("云南", "6区", "沃尔玛云南昆明万宏路分店", "昆明"),
    2507: ("四川", "6区", "沃尔玛四川绵阳绵兴东路店", "绵阳"),
    2520: ("四川", "6区", "沃尔玛四川广元东苑店（2520）", "广元"),
    2524: ("云南", "6区", "沃尔玛云南临沧沧江路分店", "临沧"),
    2550: ("四川", "6区", "沃尔玛四川自贡荣县店（2550）", "自贡"),
    2558: ("四川", "6区", "沃尔玛四川乐山峨嵋山市峨眉山店（2558）", "乐山"),
    2570: ("云南", "6区", "沃尔玛云南昆明宜良县鱼龙街分店", "昆明"),
    2573: ("四川", "6区", "沃尔玛四川成都驿都城店（2573）", "成都"),
    2593: ("四川", "6区", "沃尔玛四川成都崇州市崇州店", "成都"),
    2707: ("四川", "6区", "沃尔玛四川攀枝花清香坪店（2707）", "攀枝花"),
    2715: ("四川", "6区", "沃尔玛四川成都郫县红光店（2715店）", "成都"),
    2716: ("云南", "6区", "沃尔玛云南楚雄团结路分店", "楚雄"),
    2721: ("四川", "6区", "沃尔玛四川成都卓锦曼购中心店（2721店）", "成都"),
    2763: ("四川", "6区", "沃尔玛四川成都盛邦街店（2763）", "成都"),
    2778: ("四川", "6区", "沃尔玛四川泸州宝龙广场店", "泸州"),
    2797: ("四川", "6区", "2797成都洞子口店", "成都"),
    816: ("云南", "6区", "沃尔玛云南昆明新迎店", "昆明"),
    1021: ("云南", "6区", "沃尔玛云南玉溪东风广场店", "玉溪"),
    1092: ("云南", "6区", "沃尔玛云南昆明北辰店", "昆明"),
    1093: ("云南", "6区", "沃尔玛云南大理泰安店", "大理"),
    2030: ("云南", "6区", "沃尔玛云南昆明兴苑店", "昆明"),
    2054: ("云南", "6区", "沃尔玛云南曲靖交通路店", "曲靖"),
    2084: ("云南", "6区", "沃尔玛云南昆明云山店", "昆明"),
    2138: ("云南", "6区", "沃尔玛云南昆明前兴店", "昆明"),
    2152: ("云南", "6区", "沃尔玛云南红河个旧市个旧店", "红河"),
    2208: ("云南", "6区", "沃尔玛云南楚雄楚雄店", "楚雄"),
    2230: ("云南", "6区", "沃尔玛云南普洱林源路店", "普洱"),
    2231: ("云南", "6区", "沃尔玛云南景洪市景洪店", "西双版纳"),
    2345: ("云南", "6区", "沃尔玛云南昆明黑林铺店", "昆明"),
    2346: ("云南", "6区", "沃尔玛云南文山文山店", "文山"),
    2355: ("云南", "6区", "沃尔玛云南红河弥勒市弥勒店", "红河"),
    2385: ("云南", "6区", "沃尔玛云南昭通太平街店", "昭通"),
    2393: ("云南", "6区", "沃尔玛云南曲靖宣威市文化路店", "曲靖"),
    2464: ("云南", "6区", "沃尔玛云南曲靖子午路店", "曲靖"),
    2499: ("云南", "6区", "沃尔玛云南红河建水县建水分店", "红河"),
    2540: ("云南", "6区", "沃尔玛云南昆明人民西路分店", "昆明"),
    2572: ("云南", "6区", "沃尔玛云南丽江祥和商业广场店", "丽江"),
    2584: ("云南", "6区", "沃尔玛云南昆明呈贡店", "昆明"),
    2596: ("云南", "6区", "沃尔玛云南昆明建工新城店", "昆明"),
    2722: ("云南", "6区", "沃尔玛云南大理泰业国际广场店", "大理"),
    2732: ("云南", "6区", "沃尔玛云南昆明金瓦路分店", "昆明"),
    2767: ("云南", "6区", "沃尔玛购物广场云南曲靖罗平县罗平店", "曲靖"),
    3401: ("云南", "6区", "沃尔玛云南昆明大观店", "昆明"),
    3416: ("云南", "6区", "沃尔玛云南昆明集大店", "昆明"),
    5835: ("云南", "6区", "沃尔玛云南昆明国贸店", "昆明"),
    210: ("重庆", "6区", "沃尔玛重庆南坪天龙店（南岸区）", "重庆"),
    212: ("重庆", "6区", "沃尔玛重庆大礼堂店（渝中区）", "重庆"),
    1058: ("重庆", "6区", "沃尔玛重庆江北店（渝北区）", "重庆"),
    1075: ("重庆", "6区", "沃尔玛重庆大渡口店", "重庆"),
    2017: ("重庆", "6区", "沃尔玛重庆凤天路店(沙坪坝区)", "重庆"),
    2064: ("重庆", "6区", "沃尔玛重庆三峡福斯德广场店（万州）", "重庆"),
    2069: ("重庆", "6区", "沃尔玛重庆渝北店（渝北区）", "重庆"),
    2169: ("重庆", "6区", "沃尔玛重庆永川店", "重庆"),
    2761: ("重庆", "6区", "沃尔玛重庆回兴店", "重庆"),
    2383: ("湖北", "4区", "沃尔玛湖北荆门荆门店", "荆门"),
    2720: ("湖北", "4区", "沃尔玛湖北孝感长征路分店", "孝感"),
    2471: ("湖南", "4区", "沃尔玛湖南娄底冷水江市冷水江店", "娄底"),
    2124: ("江西", "4区", "沃尔玛江西萍乡绿茵广场店", "萍乡"),
    110: ("广东", "5区", "沃尔玛广东广州云景店", "广州"),
    605: ("福建", "5区", "沃尔玛福建福州东街店", "福州"),
    2354: ("四川", "6区", "沃尔玛四川眉山眉山店", "眉山"),
    104: ("广东", "5区", "沃尔玛广东广州广源店", "广州"),
    106: ("广东", "5区", "沃尔玛广东广州番禺店", "广州"),
    108: ("广东", "5区", "108廣州廣雅", "广州"),
    125: ("广东", "5区", "125广州黄石", "广州"),
    208: ("重庆", "6区", "沃尔玛重庆欣阳店(沙坪坝区)", "重庆"),
    209: ("四川", "6区", "209成都玉林", "成都"),
    217: ("四川", "6区", "沃尔玛四川成都都江堰市建设路店", "成都"),
    301: ("上海", "3区", "沃尔玛上海上南店", "上海"),
    303: ("江苏", "3区", "沃尔玛秦淮店", "南京"),
    307: ("江苏", "3区", "沃尔玛江苏无锡春申路店", "无锡"),
    315: ("上海", "3区", "沃尔玛上海祁连店", "上海"),
    701: ("浙江", "3区", "沃尔玛杭州黄龙店", "杭州"),
    702: ("浙江", "3区", "沃尔玛温州东海店", "温州"),
    803: ("北京", "2区", "沃尔玛北京朝阳店", "北京"),
    814: ("陕西", "2区", "814西安莲湖", "西安"),
    1005: ("重庆", "6区", "1005杨家坪", "重庆"),
    1017: ("上海", "3区", "沃尔玛上海五角场店", "上海"),
    1042: ("福建", "5区", "沃尔玛福建泉州江滨北路店", "泉州"),
    1046: ("福建", "5区", "沃尔玛福建泉州石狮市石狮店", "泉州"),
    1061: ("浙江", "3区", "沃尔玛嘉兴洪兴路店", "嘉兴"),
    1072: ("上海", "3区", "沃尔玛上海江桥店", "上海"),
    1076: ("山西", "2区", "沃尔玛山西太原三墙路店", "太原"),
    1096: ("重庆", "6区", "沃尔玛重庆北辰天街店（江北区）", "重庆"),
    2033: ("陕西", "2区", "2033西安未央路分店", "西安"),
    2039: ("江西", "4区", "沃尔玛江西南昌青山路店", "南昌"),
    2042: ("江苏", "3区", "沃尔玛江苏苏州吴江店", "苏州"),
    2060: ("四川", "6区", "沃尔玛四川绵阳站前店", "绵阳"),
    2063: ("广东", "5区", "沃尔玛广东肇庆康乐北路店", "外埠"),
    2074: ("安徽", "3区", "沃尔玛安徽阜阳人民东路店", "阜阳"),
    2083: ("广东", "5区", "沃尔玛广东惠州惠东店", "惠州"),
    2097: ("江西", "4区", "沃尔玛江西抚州赣东大道店", "抚州"),
    2114: ("北京", "2区", "沃尔玛北京清河店", "北京"),
    2118: ("福建", "5区", "沃尔玛福建泉州南安市南安店", "泉州"),
    2132: ("河南", "2区", "沃尔玛河南商丘民主店", "商丘"),
    2190: ("浙江", "3区", "沃尔玛浙江杭州富阳东方茂店", "杭州"),
    2348: ("内蒙古", "1区", "沃尔玛内蒙鄂尔多斯伊金霍洛东街分店", "鄂尔多斯"),
    2373: ("浙江", "3区", "沃尔玛嘉兴中港城店", "嘉兴"),
    2391: ("浙江", "3区", "沃尔玛浙江宁波余姚市嘉悦店", "宁波"),
    2416: ("江西", "4区", "沃尔玛江西新余巨能广场店", "新余"),
    2433: ("河南", "2区", "沃尔玛河南鹤壁鹤壁店", "鹤壁"),
    2436: ("江西", "4区", "沃尔玛江西南昌莲塘澄湖店", "南昌"),
    2459: ("浙江", "3区", "沃尔玛浙江湖州市区店", "湖州"),
    2483: ("四川", "6区", "沃尔玛四川德阳广汉市广汉店", "德阳"),
    2501: ("广东", "5区", "沃尔玛广东东莞大朗店", "东莞"),
    2525: ("江西", "4区", "沃尔玛江西宜春樟树市步行街店", "宜春"),
    2701: ("广东", "5区", "沃尔玛广东深圳洪湖店", "深圳"),
    2800: ("广东", "5区", "惠选超市新安五路分店", "深圳"),
    2801: ("广东", "5区", "沃尔玛惠选超市广东东莞红荔路分店", "东莞"),
    2803: ("广东", "5区", "沃尔玛惠选超市广东广州东莞庄路分店", "广州"),
    2804: ("广东", "5区", "沃尔玛惠选超市广东深圳海涛路分店", "深圳"),
    2805: ("广东", "5区", "深圳福田华强店", "深圳"),
    2806: ("广东", "5区", "沃尔玛惠选超市广东深圳龙岗开放路分店", "深圳"),
    2807: ("广东", "5区", "广东深圳华夏路分店", "深圳"),
    2808: ("广东", "5区", "广东东莞常平市场路分店", "东莞"),
    3424: ("广东", "5区", "沃尔玛广东深圳嘉里中心店", "深圳"),
    5017: ("JD", "JD店", "京东到家", "JD"),
    5845: ("北京", "2区", "5845北京知春路分店", "北京"),
    2818: ("广东", "5区", "深圳福田福民店", "深圳"),
    2798: ("云南", "6区", "沃尔玛云南昆明龙泉店", "昆明"),
    2799: ("云南", "6区", "沃尔玛云南昆明正大店", "昆明"),
}

def get_store_info(club_nbr):
    """返回 (省份, 大区, 门店名称, 城市)"""
    info = STORE_INFO.get(club_nbr)
    if info:
        return info
    return ("未知", "未知", "未知门店", "未知")

# ==================== 解析订单文件 ====================
def parse_order_metadata(filepath):
    """读取表头行，识别各 Range 的 Qty 和 Sales 列名，以及时间范围"""
    df_raw = pd.read_excel(filepath, header=None, nrows=100, dtype=str)
    range_dates = {}
    header_row = None
    date_pattern = re.compile(
        r"Time Range (\d+) Is Between\s+(\d{2})-(\d{2})-(\d{4})\s+and\s+(\d{2})-(\d{2})-(\d{4})",
        re.IGNORECASE
    )
    for idx, row in df_raw.iterrows():
        if row.isnull().all():
            continue
        row_str = " ".join([str(v) for v in row if pd.notna(v)])
        match = date_pattern.search(row_str)
        if match:
            rn = int(match.group(1))
            start_month = int(match.group(2))
            start_day = int(match.group(3))
            start_year = int(match.group(4))
            end_month = int(match.group(5))
            end_day = int(match.group(6))
            end_year = int(match.group(7))
            try:
                start_date = datetime(start_year, start_month, start_day)
                end_date = datetime(end_year, end_month, end_day)
                range_dates[rn] = {"start": start_date, "end": end_date}
            except ValueError as e:
                raise ValueError(f"日期解析错误 Range {rn}: {e}")
        # 寻找表头行：包含 "UPC" 和 "Store Nbr"
        if "UPC" in row.values and "Store Nbr" in row.values:
            header_row = idx
            break
    if header_row is None:
        raise ValueError("未找到数据表头行（缺少 UPC 或 Store Nbr）")
    # 读取全表
    df_full = pd.read_excel(filepath, header=header_row, dtype=str)
    # 识别销量、销售额、库存、在途列
    range_qty_cols = {}
    range_sales_cols = {}
    range_oh_cols = {}  # 库存
    range_oo_cols = {}  # 在途
    for col in df_full.columns:
        qty_match = re.match(r"Range\s*(\d+)\s*POS\s*Qty", col, re.IGNORECASE)
        sales_match = re.match(r"Range\s*(\d+)\s*POS\s*Sales", col, re.IGNORECASE)
        oh_match = re.match(r"Range\s*(\d+)\s*(?:Current\s*)?On-Hand\s*Qty", col, re.IGNORECASE)
        oo_match = re.match(r"Range\s*(\d+)\s*(?:Current\s*)?On-Order\s*Qty", col, re.IGNORECASE)
        if qty_match:
            rn = int(qty_match.group(1))
            range_qty_cols[rn] = col
        if sales_match:
            rn = int(sales_match.group(1))
            range_sales_cols[rn] = col
        if oh_match:
            rn = int(oh_match.group(1))
            range_oh_cols[rn] = col
        if oo_match:
            rn = int(oo_match.group(1))
            range_oo_cols[rn] = col
    if not range_qty_cols:
        raise ValueError("未找到 Range N POS Qty 列（销量）")
    if not range_sales_cols:
        raise ValueError("未找到 Range N POS Sales 列（销额）")
    return range_dates, header_row, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols

def load_and_aggregate(filepath, header_row, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols):
    """读取数据并返回三个聚合 DataFrame：门店、品项、省份（原始分组）"""
    df = pd.read_excel(filepath, header=header_row, dtype=str)
    # 必要列检查
    if "UPC" not in df.columns or "Store Nbr" not in df.columns:
        raise ValueError("文件缺少 UPC 或 Store Nbr 列")
    df["UPC"] = df["UPC"].astype(str).str.strip()
    df["Store Nbr"] = df["Store Nbr"].astype(str).str.extract(r"(\d+)")[0].astype(float).astype(int)

    # 处理销量、销售额、库存和在途列
    all_range_nums = sorted(set(list(range_qty_cols.keys()) + list(range_sales_cols.keys())))
    for rn in all_range_nums:
        if rn in range_qty_cols:
            col = range_qty_cols[rn]
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        if rn in range_sales_cols:
            col = range_sales_cols[rn]
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        # 只有Range1才处理库存和在途
        if rn == 1:
            if rn in range_oh_cols:
                col = range_oh_cols[rn]
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            if rn in range_oo_cols:
                col = range_oo_cols[rn]
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # 品项聚合（按 UPC）
    item_agg = {}
    for rn in all_range_nums:
        if rn in range_qty_cols:
            item_agg[range_qty_cols[rn]] = "sum"
        if rn in range_sales_cols:
            item_agg[range_sales_cols[rn]] = "sum"
        # 只有Range1添加库存和在途
        if rn == 1:
            if rn in range_oh_cols:
                item_agg[range_oh_cols[rn]] = "sum"
            if rn in range_oo_cols:
                item_agg[range_oo_cols[rn]] = "sum"
    if "Item Desc 1" in df.columns:
        item_group = df.groupby("UPC", as_index=False).agg(
            {**item_agg, "Item Desc 1": "first"}
        )
    else:
        item_group = df.groupby("UPC", as_index=False).agg(item_agg)
        item_group["Item Desc 1"] = ""

    # 门店聚合
    store_agg = item_agg.copy()
    store_group = df.groupby("Store Nbr", as_index=False).agg(store_agg)

    # 添加门店信息
    store_info = store_group["Store Nbr"].apply(lambda x: get_store_info(x))
    store_group["省份"] = store_info.apply(lambda x: x[0])
    store_group["大区"] = store_info.apply(lambda x: x[1])
    store_group["门店名称"] = store_info.apply(lambda x: x[2])
    store_group["城市"] = store_info.apply(lambda x: x[3])

    # 省份聚合（基于门店聚合结果）
    prov_group = store_group.groupby("省份", as_index=False).agg(
        {col: "sum" for col in store_agg.keys()}
    )

    # --- 关键修正：显式创建副本，消除 SettingWithCopyWarning ---
    prov_df = prov_group.copy()   # <--- 添加 .copy()

    return all_range_nums, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols, item_group, store_group, prov_df, df
    
# ==================== 成长率计算 ====================
def growth_rate(current, previous):
    if previous == 0:
        return float('inf') if current > 0 else 0
    return (current - previous) / previous * 100

# ==================== 导出 Excel（含品项sheet） ====================
def export_optimized_excel(order_file, output_path):
    try:
        range_dates, header_row, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols = parse_order_metadata(order_file)
        all_ranges, qty_cols, sales_cols, oh_cols, oo_cols, item_df, store_df, prov_df, _ = load_and_aggregate(
            order_file, header_row, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols
        )

        # ------------------ 统一重命名列为中文 ------------------
        rename_dict = {}
        for rn in all_ranges:
            if rn in qty_cols:
                rename_dict[qty_cols[rn]] = f"Range{rn}_销量"
            if rn in sales_cols:
                rename_dict[sales_cols[rn]] = f"Range{rn}_销售额"
            # 只有Range1有库存和在途
            if rn == 1:
                if rn in oh_cols:
                    rename_dict[oh_cols[rn]] = f"Range{rn}_库存"
                if rn in oo_cols:
                    rename_dict[oo_cols[rn]] = f"Range{rn}_在途"
        item_df.rename(columns=rename_dict, inplace=True)
        store_df.rename(columns=rename_dict, inplace=True)
        prov_df.rename(columns=rename_dict, inplace=True)

        # ------------------ 品项汇总表 ------------------
        def make_item_name(row):
            upc = row["UPC"]
            desc = row.get("Item Desc 1", "")
            if pd.notna(desc) and desc:
                return f"{desc} ({upc})"
            else:
                return f"商品{upc}"
        item_df["品名"] = item_df.apply(make_item_name, axis=1)
        base_item_cols = ["品名", "UPC", "Item Desc 1"]
        item_range_cols = [c for c in item_df.columns if c.startswith("Range") and "成长率" not in c]
        # 自定义排序：先按 Range数字，再按指标类型（销量 < 销售额 < 库存 < 在途）
        def sort_key(x):
            m = re.search(r"Range(\d+)_(.*)", x)
            if m:
                rn = int(m.group(1))
                metric = m.group(2)
                metric_order = {"销量": 1, "销售额": 2, "库存": 3, "在途": 4}
                return (rn, metric_order.get(metric, 99))
            return (999, x)
        item_range_cols.sort(key=sort_key)
        item_df = item_df[base_item_cols + item_range_cols]

        # ------------------ 门店汇总表 ------------------
        # 添加成长率（销量 & 销售额）
        for df in [store_df, prov_df]:
            for prefix in ["销量", "销售额"]:
                if f"Range2_{prefix}" in df.columns and f"Range3_{prefix}" in df.columns:
                    df[f"Range2-3_{prefix}成长率"] = df.apply(
                        lambda row: growth_rate(row[f"Range2_{prefix}"], row[f"Range3_{prefix}"]), axis=1)
                if f"Range4_{prefix}" in df.columns and f"Range5_{prefix}" in df.columns:
                    df[f"Range4-5_{prefix}成长率"] = df.apply(
                        lambda row: growth_rate(row[f"Range4_{prefix}"], row[f"Range5_{prefix}"]), axis=1)

        # 门店信息列
        store_base = ["Store Nbr", "门店名称", "省份", "大区", "城市"]
        store_range_cols = [c for c in store_df.columns if c.startswith("Range") and "成长率" not in c]
        store_range_cols.sort(key=sort_key)
        store_growth_cols = [c for c in store_df.columns if "成长率" in c]
        store_df = store_df[store_base + store_range_cols + store_growth_cols]
        store_df.sort_values("Range1_销量", ascending=False, inplace=True)

        # ------------------ 省份汇总表 ------------------
        # 添加省份下最好/最差门店（基于 Range1 销量）
        if "Range1_销量" in store_df.columns:
            prov_best = {}
            prov_worst = {}
            for prov, sub_df in store_df.groupby("省份"):
                if not sub_df.empty:
                    best_row = sub_df.loc[sub_df["Range1_销量"].idxmax()]
                    worst_row = sub_df.loc[sub_df["Range1_销量"].idxmin()]
                    prov_best[prov] = f"{best_row['门店名称']} ({int(best_row['Range1_销量'])}瓶)"
                    prov_worst[prov] = f"{worst_row['门店名称']} ({int(worst_row['Range1_销量'])}瓶)"
            prov_df["销量最好门店"] = prov_df["省份"].map(prov_best).fillna("")
            prov_df["销量最差门店"] = prov_df["省份"].map(prov_worst).fillna("")

        prov_base = ["省份"]
        prov_range_cols = [c for c in prov_df.columns if c.startswith("Range") and "成长率" not in c]
        prov_range_cols.sort(key=sort_key)
        prov_growth_cols = [c for c in prov_df.columns if "成长率" in c]
        prov_extra = ["销量最好门店", "销量最差门店"]
        prov_df = prov_df[prov_base + prov_range_cols + prov_growth_cols + prov_extra]

        # ---------- 关键修正：改用非原位排序，避免 SettingWithCopyWarning ----------
        prov_df = prov_df.sort_values("Range1_销量", ascending=False)

        # ------------------ 写入 Excel ------------------
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            item_df.to_excel(writer, sheet_name='品项汇总', index=False, startrow=1)
            store_df.to_excel(writer, sheet_name='门店汇总', index=False, startrow=1)
            prov_df.to_excel(writer, sheet_name='省份汇总', index=False, startrow=1)

            # ------------------ 格式化函数（双行表头） ------------------
            def create_multi_header(ws, df, start_row=1):
                # 找出基础列和 Range 列的分界
                base_cols_end = 0
                for i, col in enumerate(df.columns):
                    if col.startswith("Range") or "成长率" in col or col in ["销量最好门店", "销量最差门店"]:
                        base_cols_end = i
                        break
                    else:
                        base_cols_end = i + 1
                # 第一行基础列留空
                for col_idx in range(1, base_cols_end + 1):
                    ws.cell(row=start_row, column=col_idx, value="")
                    ws.cell(row=start_row, column=col_idx).font = Font(bold=True)
                # 收集 Range 数字
                range_nums = set()
                for col in df.columns[base_cols_end:]:
                    if col.startswith("Range") and "成长率" not in col:
                        m = re.search(r"Range(\d+)", col)
                        if m:
                            range_nums.add(int(m.group(1)))
                sorted_ranges = sorted(range_nums)
                col_groups = {}
                current_col = base_cols_end + 1
                for rn in sorted_ranges:
                    start_col = current_col
                    for col in df.columns[base_cols_end:]:
                        if col.startswith(f"Range{rn}_"):
                            current_col += 1
                    col_groups[f"Range{rn}"] = (start_col, current_col - 1)
                # 成长率分组
                growth_start = current_col
                for col in df.columns[base_cols_end:]:
                    if "成长率" in col:
                        current_col += 1
                if growth_start < current_col:
                    col_groups["成长率"] = (growth_start, current_col - 1)
                # 额外列（销量最好/最差门店）
                extra_start = current_col
                for col in df.columns[base_cols_end:]:
                    if col in ["销量最好门店", "销量最差门店"]:
                        current_col += 1
                if extra_start < current_col:
                    col_groups["门店表现"] = (extra_start, current_col - 1)
                # 合并单元格并写入分组名
                for group_name, (start, end) in col_groups.items():
                    if start == end:
                        ws.cell(row=start_row, column=start, value=group_name)
                        ws.cell(row=start_row, column=start).font = Font(bold=True)
                        ws.cell(row=start_row, column=start).alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        ws.merge_cells(start_row=start_row, start_column=start, end_row=start_row, end_column=end)
                        ws.cell(row=start_row, column=start, value=group_name)
                        ws.cell(row=start_row, column=start).font = Font(bold=True)
                        ws.cell(row=start_row, column=start).alignment = Alignment(horizontal='center', vertical='center')

            def format_sheet(ws, df, sheet_name):
                # 写双行表头
                create_multi_header(ws, df, start_row=1)
                # 第二行列名加粗居中
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # 列宽 & 对齐
                for col_idx, col_name in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for row in range(2, ws.max_row + 1):
                        cell_val = ws.cell(row=row, column=col_idx).value
                        if cell_val:
                            cell_str = str(cell_val)
                            width = 0
                            for ch in cell_str:
                                if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                                    width += 2
                                else:
                                    width += 1
                            if width > max_len:
                                max_len = width
                    adjusted_width = min(max(max_len + 2, 8), 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                    # 数值右对齐（包括库存、在途）
                    align = 'right' if any(kw in col_name for kw in ['销量', '销售额', '库存', '在途', '成长率', 'Nbr']) else 'left'
                    for row in range(3, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.alignment = Alignment(horizontal=align, vertical='center')
                        if "成长率" in col_name and cell.value is not None:
                            if isinstance(cell.value, (int, float)):
                                if cell.value == float('inf'):
                                    cell.value = "∞"
                                else:
                                    cell.value = cell.value / 100  # 转换为0-1范围的小数
                                    cell.number_format = '0.00%'  # 设置为百分比格式

            # 格式化三个 sheet
            for sheet_name, df in [('品项汇总', item_df), ('门店汇总', store_df), ('省份汇总', prov_df)]:
                ws = writer.sheets[sheet_name]
                format_sheet(ws, df, sheet_name)

                # 添加总计行（所有 sheet）
                base_cols = []
                if sheet_name == '品项汇总':
                    base_cols = ["品名", "UPC", "Item Desc 1"]
                elif sheet_name == '门店汇总':
                    base_cols = ["Store Nbr", "门店名称", "省份", "大区", "城市"]
                else:
                    base_cols = ["省份"]
                numeric_cols = [c for c in df.columns if c not in base_cols and "成长率" not in c and c not in ["销量最好门店", "销量最差门店"]]
                # 重要：先记录数据行数，再添加总计行
                data_row_count = ws.max_row
                sum_row = data_row_count + 1
                ws.cell(row=sum_row, column=1, value="总计")
                ws.cell(row=sum_row, column=1).font = Font(bold=True)
                total_col_indices = {}
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if col_name in numeric_cols:
                        # 使用 Excel SUM 公式
                        col_letter = get_column_letter(col_idx)
                        formula = f"=SUM({col_letter}3:{col_letter}{data_row_count})"
                        ws.cell(row=sum_row, column=col_idx, value=formula)
                        ws.cell(row=sum_row, column=col_idx).font = Font(bold=True)
                        ws.cell(row=sum_row, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')
                        total_col_indices[col_name] = col_idx
                    # 品项汇总的所有列左对齐
                    if sheet_name == '品项汇总':
                        for row in range(3, data_row_count + 1):
                            ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='left', vertical='center')
                # 计算成长率总计（也使用公式）
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if "成长率" in col_name:
                        if "Range2-3" in col_name:
                            if "销量" in col_name:
                                cur_col_idx = total_col_indices.get("Range2_销量")
                                prev_col_idx = total_col_indices.get("Range3_销量")
                            else:
                                cur_col_idx = total_col_indices.get("Range2_销售额")
                                prev_col_idx = total_col_indices.get("Range3_销售额")
                        elif "Range4-5" in col_name:
                            if "销量" in col_name:
                                cur_col_idx = total_col_indices.get("Range4_销量")
                                prev_col_idx = total_col_indices.get("Range5_销量")
                            else:
                                cur_col_idx = total_col_indices.get("Range4_销售额")
                                prev_col_idx = total_col_indices.get("Range5_销售额")
                        else:
                            continue
                        if cur_col_idx and prev_col_idx:
                            cur_letter = get_column_letter(cur_col_idx)
                            prev_letter = get_column_letter(prev_col_idx)
                            # 使用Excel公式并设置单元格格式为百分比
                            formula = f'=IF({prev_letter}{sum_row}=0,IF({cur_letter}{sum_row}>0,"∞",0),(({cur_letter}{sum_row}-{prev_letter}{sum_row})/{prev_letter}{sum_row}))'
                            cell = ws.cell(row=sum_row, column=col_idx, value=formula)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            # 设置百分比格式，两位小数
                            if cell.value != "∞":
                                cell.number_format = '0.00%'

            # ------------------ 时间范围说明 Sheet ------------------
            if range_dates:
                meta_df = pd.DataFrame([
                    {"时间段": f"Range{rn}", "开始日期": info["start"].strftime("%Y-%m-%d"), "结束日期": info["end"].strftime("%Y-%m-%d")}
                    for rn, info in range_dates.items()
                ])
                meta_df.to_excel(writer, sheet_name="时间范围说明", index=False)
                ws_meta = writer.sheets["时间范围说明"]
                for cell in ws_meta[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for col_idx, col_name in enumerate(meta_df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for cell in ws_meta[col_letter]:
                        if cell.value:
                            cell_str = str(cell.value)
                            width = 0
                            for ch in cell_str:
                                if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                                    width += 2
                                else:
                                    width += 1
                            if width > max_len:
                                max_len = width
                    adjusted_width = min(max(max_len + 2, 10), 50)
                    ws_meta.column_dimensions[col_letter].width = adjusted_width
                    for row in range(2, ws_meta.max_row + 1):
                        ws_meta.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

        return True, None
    except Exception as e:
        return False, str(e)
        
# ==================== 文本分析（控制台输出） ====================
def analyze_sales(order_file):
    try:
        range_dates, header_row, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols = parse_order_metadata(order_file)
        all_ranges, qty_cols, sales_cols, oh_cols, oo_cols, item_df, store_df, prov_df, _ = load_and_aggregate(
            order_file, header_row, range_qty_cols, range_sales_cols, range_oh_cols, range_oo_cols
        )

        # ---------- 重命名 item_df 的列为中文（与导出 Excel 保持一致） ----------
        rename_dict = {}
        for rn in all_ranges:
            if rn in qty_cols:
                rename_dict[qty_cols[rn]] = f"Range{rn}_销量"
            if rn in sales_cols:
                rename_dict[sales_cols[rn]] = f"Range{rn}_销售额"
            # 只有Range1有库存和在途
            if rn == 1:
                if rn in oh_cols:
                    rename_dict[oh_cols[rn]] = f"Range{rn}_库存"
                if rn in oo_cols:
                    rename_dict[oo_cols[rn]] = f"Range{rn}_在途"
        item_df.rename(columns=rename_dict, inplace=True)

        # 添加品名（合并 Item Desc 1 和 UPC）
        def make_item_name(row):
            upc = row["UPC"]
            desc = row.get("Item Desc 1", "")
            if pd.notna(desc) and desc:
                return f"{desc} ({upc})"
            else:
                return f"商品{upc}"
        item_df["品名"] = item_df.apply(make_item_name, axis=1)

        # ---------- 构建文本表格（销量 / 销额） ----------
        def build_text_table(df, metric_type):
            is_qty = (metric_type == 'Qty')
            prefix = "销量" if is_qty else "销售额"
            # 找出所有 RangeX_销量/销售额 列
            col_map = {}
            for rn in all_ranges:
                col_name = f"Range{rn}_{prefix}"
                if col_name in df.columns:
                    col_map[rn] = col_name
            if not col_map:
                return "无数据"

            headers = ["品名", "UPC"]
            if 1 in col_map: headers.append(f"日{prefix}")
            if 2 in col_map: headers.append(f"月{prefix}")
            if 3 in col_map: headers.append(f"去年同月{prefix}")
            if 2 in col_map and 3 in col_map: headers.append(f"{prefix}月成长率")
            if 4 in col_map: headers.append(f"今年至今{prefix}")
            if 5 in col_map: headers.append(f"去年同期{prefix}")
            if 4 in col_map and 5 in col_map: headers.append(f"{prefix}YTD成长率")

            rows = []
            for _, row in df.iterrows():
                line = [row["品名"], row["UPC"]]
                if 1 in col_map:
                    val = row[col_map[1]]
                    line.append(f"{val:,.0f}" if is_qty else f"{val:,.2f}")
                if 2 in col_map:
                    val = row[col_map[2]]
                    line.append(f"{val:,.0f}" if is_qty else f"{val:,.2f}")
                if 3 in col_map:
                    val = row[col_map[3]]
                    line.append(f"{val:,.0f}" if is_qty else f"{val:,.2f}")
                if 2 in col_map and 3 in col_map:
                    cur = row[col_map[2]]
                    prev = row[col_map[3]]
                    g = growth_rate(cur, prev)
                    line.append(f"{g:.1f}%" if g != float('inf') else "∞")
                if 4 in col_map:
                    val = row[col_map[4]]
                    line.append(f"{val:,.0f}" if is_qty else f"{val:,.2f}")
                if 5 in col_map:
                    val = row[col_map[5]]
                    line.append(f"{val:,.0f}" if is_qty else f"{val:,.2f}")
                if 4 in col_map and 5 in col_map:
                    cur = row[col_map[4]]
                    prev = row[col_map[5]]
                    g = growth_rate(cur, prev)
                    line.append(f"{g:.1f}%" if g != float('inf') else "∞")
                rows.append(line)

            # 总计行
            total_line = ["【总计】", ""]
            for rn in all_ranges:
                if rn in col_map:
                    total = df[col_map[rn]].sum()
                    total_line.append(f"{total:,.0f}" if is_qty else f"{total:,.2f}")

            output = []
            output.append("   ".join(headers))
            output.append("-" * 80)
            for r in rows:
                output.append("   ".join(str(x) for x in r))
            output.append("-" * 80)
            output.append("   ".join(str(x) for x in total_line))
            return "\n".join(output)

        result = []
        result.append("")
        result.append("=" * 80)
        result.append(f"订单文件：{order_file}")
        if 1 in range_dates:
            result.append(f"销售日期：{range_dates[1]['end'].strftime('%Y年%m月%d日')}")
        result.append("=" * 80)
        result.append("")
        result.append("【销量分析（单位：瓶）】")
        result.append(build_text_table(item_df, 'Qty'))
        result.append("")
        result.append("【销额分析（单位：元）】")
        result.append(build_text_table(item_df, 'Sales'))
        result.append("")
        result.append("【各时间段说明】")
        for rn in sorted(range_dates.keys()):
            start = range_dates[rn]["start"].strftime("%Y-%m-%d")
            end = range_dates[rn]["end"].strftime("%Y-%m-%d")
            result.append(f"  Range{rn}：{start} 至 {end}")
        return "\n".join(result), None
    except Exception as e:
        return None, str(e)

# ==================== GUI ====================
class SalesAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("hefengfan沃尔玛销售分析")
        self.root.geometry("500x600")
        self.root.option_add('*Font', ('微软雅黑', 10))

        self.order_path = tk.StringVar()
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self.root, text="订单文件（必选）:", font=("微软雅黑", 10)).pack(pady=(10,0), anchor="w", padx=10)
        frame_order = tk.Frame(self.root)
        frame_order.pack(fill="x", padx=10, pady=5)
        tk.Entry(frame_order, textvariable=self.order_path, width=40, font=("微软雅黑", 10)).pack(side="left", fill="x", expand=True)
        tk.Button(frame_order, text="浏览...", command=self.browse_order).pack(side="right", padx=5)

        self.analyze_btn = tk.Button(self.root, text="开始分析", command=self.analyze, width=20, font=("微软雅黑", 10))
        self.analyze_btn.pack(pady=10)

        self.export_btn = tk.Button(self.root, text="导出优化 Excel", command=self.export_excel, width=20, font=("微软雅黑", 10))
        self.export_btn.pack(pady=5)

        self.result_text = scrolledtext.ScrolledText(self.root, wrap=tk.NONE, width=70, height=30, font=("Consolas", 10))
        self.result_text.pack(padx=10, pady=5, fill="both", expand=True)

    def browse_order(self):
        fname = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
        if fname:
            self.order_path.set(fname)

    def analyze(self):
        order_file = self.order_path.get().strip()
        if not order_file:
            messagebox.showerror("错误", "请先选择订单文件")
            return

        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, "正在分析，请稍候...\n")
        self.root.update()

        result, error = analyze_sales(order_file)
        if error:
            messagebox.showerror("分析失败", error)
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, f"发生错误：\n{error}")
        else:
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, result)

    def export_excel(self):
        order_file = self.order_path.get().strip()
        if not order_file:
            messagebox.showerror("错误", "请先选择订单文件")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="保存优化Excel报表"
        )
        if not output_path:
            return

        self.result_text.insert(tk.END, "\n正在导出优化Excel，请稍候...\n")
        self.root.update()

        success, error = export_optimized_excel(order_file, output_path)
        if success:
            messagebox.showinfo("导出成功", f"报表已保存至：\n{output_path}")
            self.result_text.insert(tk.END, f"\n✅ 优化Excel导出成功：{output_path}\n")
        else:
            messagebox.showerror("导出失败", error)
            self.result_text.insert(tk.END, f"\n❌ 导出失败：{error}\n")

if __name__ == "__main__":
    root = tk.Tk()
    app = SalesAnalyzerGUI(root)
    root.mainloop()
