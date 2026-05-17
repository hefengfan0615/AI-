import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import pandas as pd
import re
from datetime import datetime
import warnings
import locale
import numpy as np
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------- 1. 抑制 OpenPyXL 警告 ----------
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# ---------- 2. 修正 Tkinter 中文报错 ----------
try:
    locale.setlocale(locale.LC_CTYPE, 'Chinese')
except:
    try:
        locale.setlocale(locale.LC_CTYPE, 'zh_CN.UTF-8')
    except:
        pass

# ==================== 内置门店字典 ====================
BUILTIN_STORE_DICT = {
    4754: {"type": "门店", "province": "福建", "erp": "山姆沃尔玛福建福州鼓楼店"},
    4788: {"type": "门店", "province": "北京", "erp": "山姆沃尔玛北京石景山店"},
    4801: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东广州番禺店"},
    4803: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东深圳龙岗店"},
    4805: {"type": "门店", "province": "北京", "erp": "山姆沃尔玛北京亦庄店"},
    4806: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏常州新北店"},
    4807: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海浦东店"},
    4808: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江杭州西溪店"},
    4809: {"type": "门店", "province": "辽宁", "erp": "山姆沃尔玛辽宁大连香炉礁店"},
    4814: {"type": "门店", "province": "湖北", "erp": "山姆沃尔玛湖北武汉汉口店"},
    4818: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏苏州金鸡湖店"},
    4829: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东珠海香洲店"},
    4830: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东深圳龙华店"},
    4832: {"type": "门店", "province": "江西", "erp": "山姆沃尔玛江西南昌红谷滩店"},
    4833: {"type": "门店", "province": "辽宁", "erp": "山姆沃尔玛辽宁沈阳和平店"},
    4834: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏苏州木渎店"},
    4835: {"type": "门店", "province": "天津", "erp": "山姆沃尔玛天津梅江店"},
    4836: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏南京雨花台店"},
    4839: {"type": "门店", "province": "四川", "erp": "山姆沃尔玛四川成都华侨城店"},
    4840: {"type": "门店", "province": "湖南", "erp": "山姆沃尔玛湖南长沙雨花店"},
    4841: {"type": "门店", "province": "福建", "erp": "山姆沃尔玛福建厦门自贸区店"},
    4846: {"type": "门店", "province": "北京", "erp": "山姆沃尔玛北京顺义店"},
    4851: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江宁波江北店"},
    4852: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏南通崇川店"},
    4865: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海青浦店"},
    6058: {"type": "前置仓", "province": "广东", "erp": "山姆沃尔玛广东珠海前置仓"},
    6158: {"type": "前置仓", "province": "广东", "erp": "山姆沃尔玛广东广州前置仓"},
    6258: {"type": "前置仓", "province": "北京", "erp": "山姆沃尔玛广东惠州6258前置仓"},
    6259: {"type": "前置仓", "province": "天津", "erp": "山姆沃尔玛广东惠州6259前置仓"},
    6505: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东深圳福田店"},
    6507: {"type": "门店", "province": "福建", "erp": "山姆沃尔玛福建福州仓山店"},
    6508: {"type": "门店", "province": "重庆", "erp": "山姆沃尔玛重庆九龙坡店"},
    6509: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏苏州昆山店"},
    6510: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏南京江北店"},
    6511: {"type": "门店", "province": "四川", "erp": "山姆沃尔玛四川成都店"},
    6512: {"type": "门店", "province": "湖北", "erp": "山姆沃尔玛湖北武汉光谷店"},
    6513: {"type": "门店", "province": "重庆", "erp": "山姆沃尔玛重庆两江新区店"},
    6515: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江杭州萧山店"},
    6516: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东惠州店"},
    6517: {"type": "门店", "province": "湖南", "erp": "山姆沃尔玛湖南长沙岳麓店"},
    6518: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东广州东店"},
    6519: {"type": "门店", "province": "北京", "erp": "山姆沃尔玛北京大兴店"},
    6523: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江杭州北店"},
    6526: {"type": "门店", "province": "湖北", "erp": "山姆沃尔玛湖北武汉江岸后湖店"},
    6529: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江宁波鄞州店"},
    6532: {"type": "门店", "province": "广西", "erp": "山姆沃尔玛广西南宁店"},
    6533: {"type": "门店", "province": "四川", "erp": "山姆沃尔玛四川成都武侯店"},
    6538: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海嘉定店"},
    6551: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东深圳宝安店"},
    6557: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江绍兴店"},
    6558: {"type": "前置仓", "province": "上海", "erp": "山姆沃尔玛上海全渠道履约中心"},
    6559: {"type": "履约中心", "province": "江苏", "erp": "山姆沃尔玛广东惠州履约中心6559"},
    6560: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏无锡店"},
    6567: {"type": "门店", "province": "湖北", "erp": "山姆沃尔玛湖北武汉汉阳店"},
    6568: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海真如店"},
    6569: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江嘉兴南湖店"},
    6570: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东东莞店"},
    6571: {"type": "门店", "province": "福建", "erp": "山姆沃尔玛福建泉州晋江店"},
    6580: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东深圳前海店"},
    6582: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东佛山顺德店"},
    6583: {"type": "门店", "province": "浙江", "erp": "山姆沃尔玛浙江温州店"},
    6587: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海浦东东店"},
    6590: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东广州荔湾店"},
    6599: {"type": "门店", "province": "北京", "erp": "山姆沃尔玛北京昌平店"},
    6600: {"type": "前置仓", "province": "天津", "erp": "山姆沃尔玛广东惠州6600前置仓"},
    6601: {"type": "前置仓", "province": "天津", "erp": "山姆沃尔玛广东惠州6601前置仓"},
    6602: {"type": "前置仓", "province": "广东", "erp": "山姆沃尔玛广东广州6602前置仓（次级）"},
    6660: {"type": "京东VC", "province": "广东", "erp": "山姆沃尔玛广东广州6660京东仓"},
    6661: {"type": "京东VC", "province": "辽宁", "erp": "山姆沃尔玛广东广州6661京东仓"},
    6662: {"type": "京东VC", "province": "湖北", "erp": "山姆沃尔玛广东广州6662京东仓"},
    6663: {"type": "履约中心", "province": "天津", "erp": "山姆沃尔玛北京虚拟店"},
    6664: {"type": "京东VC", "province": "四川", "erp": "山姆沃尔玛广东广州6664京东仓"},
    6665: {"type": "京东VC", "province": "上海", "erp": "山姆沃尔玛广东广州6665京东仓"},
    6668: {"type": "京东VC", "province": "北京", "erp": "山姆沃尔玛广东广州6668京东仓"},
    6682: {"type": "前置仓", "province": "上海", "erp": "山姆沃尔玛上海云仓"},
    6758: {"type": "履约中心", "province": "广东", "erp": "山姆沃尔玛广东深圳电商拣货中心"},
    6852: {"type": "履约中心", "province": "辽宁", "erp": "山姆沃尔玛广东惠州6852履约中心"},
    6855: {"type": "履约中心", "province": "湖南", "erp": "山姆沃尔玛广东惠州6855履约中心"},
    6856: {"type": "履约中心", "province": "四川", "erp": "山姆沃尔玛广东惠州西南履约中心"},
    6858: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海宝山店"},
    6865: {"type": "前置仓", "province": "辽宁", "erp": "山姆沃尔玛广东惠州6865前置仓"},
    6866: {"type": "前置仓", "province": "广东", "erp": "山姆沃尔玛广东惠州6866前置仓"},
    6886: {"type": "前置仓", "province": "江西", "erp": "山姆沃尔玛广东惠州6886前置仓"},
    6887: {"type": "前置仓", "province": "辽宁", "erp": "山姆沃尔玛广东惠州6887前置仓"},
    6902: {"type": "门店", "province": "安徽", "erp": "山姆沃尔玛安徽合肥店"},
    6905: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏扬州店"},
    6908: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏张家港店"},
    6909: {"type": "门店", "province": "江苏", "erp": "山姆沃尔玛江苏无锡惠山店"},
    6911: {"type": "门店", "province": "广东", "erp": "山姆沃尔玛广东中山石歧店"},
    6912: {"type": "门店", "province": "山东", "erp": "山姆沃尔玛山东青岛店"},
    6917: {"type": "门店", "province": "山东", "erp": "山姆沃尔玛山东济南店"},
    6988: {"type": "门店", "province": "上海", "erp": "山姆沃尔玛上海外高桥店"},
}

def get_store_info(club_nbr):
    info = BUILTIN_STORE_DICT.get(club_nbr)
    if info:
        return (info["type"], info["province"], info["erp"])
    return ("未知", "未知", "未知")

# ---------- 文本表格格式化模块 ----------
def str_display_width(text):
    """返回字符串的显示宽度（英文数字=1，中文=2）"""
    width = 0
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
            width += 2
        else:
            width += 1
    return width

def fit_cell_by_width(text, max_width, ellipsis='…'):
    """根据显示宽度截断文本，尾部加省略号"""
    if str_display_width(text) <= max_width:
        return text
    truncated = ''
    cur_width = 0
    for ch in text:
        ch_width = 2 if ('\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef') else 1
        if cur_width + ch_width <= max_width - str_display_width(ellipsis):
            truncated += ch
            cur_width += ch_width
        else:
            break
    return truncated + ellipsis

def create_table(headers, rows, col_align=None, max_col_width=30):
    if not rows:
        return "（无数据）\n"
    if col_align is None:
        col_align = ['left'] * len(headers)

    col_widths = []
    for i, header in enumerate(headers):
        max_w = str_display_width(header)
        for row in rows:
            cell = str(row[i]) if i < len(row) else ''
            max_w = max(max_w, str_display_width(cell))
        col_widths.append(min(max_w, max_col_width))

    sep = '+' + '+'.join(['-' * (w + 2) for w in col_widths]) + '+'

    header_cells = []
    for i, h in enumerate(headers):
        disp = fit_cell_by_width(h, col_widths[i])
        if col_align[i] == 'right':
            header_cells.append(disp.rjust(col_widths[i]))
        else:
            header_cells.append(disp.ljust(col_widths[i]))
    header_line = '| ' + ' | '.join(header_cells) + ' |'

    lines = [sep, header_line, sep]
    for row in rows:
        cells = []
        for i, val in enumerate(row):
            cell_str = str(val)
            disp = fit_cell_by_width(cell_str, col_widths[i])
            if col_align[i] == 'right':
                cells.append(disp.rjust(col_widths[i]))
            else:
                cells.append(disp.ljust(col_widths[i]))
        lines.append('| ' + ' | '.join(cells) + ' |')
    lines.append(sep)
    return '\n'.join(lines)

def parse_order_metadata(filepath):
    df_raw = pd.read_excel(filepath, header=None, nrows=50, dtype=str)
    range_dates = {}
    header_row = None
    date_pattern = re.compile(
        r"Time Range (\d+) Is Between\s+(\d{2})-(\d{2})-(\d{4})\s+and\s+(\d{2})-(\d{2})-(\d{4})",
        re.IGNORECASE
    )
    for idx, row in df_raw.iterrows():
        if row.isnull().all():
            continue
        row_str = " ".join([str(v) for v in row if pd.notna(v)])
        match = date_pattern.search(row_str)
        if match:
            rn = int(match.group(1))
            start_month = int(match.group(2))
            start_day = int(match.group(3))
            start_year = int(match.group(4))
            end_month = int(match.group(5))
            end_day = int(match.group(6))
            end_year = int(match.group(7))
            try:
                start_date = datetime(start_year, start_month, start_day)
                end_date = datetime(end_year, end_month, end_day)
                range_dates[rn] = {"start": start_date, "end": end_date}
            except ValueError as e:
                raise ValueError(f"日期解析错误 Range {rn}: {e}")
        if row_str.strip().startswith("Item Nbr"):
            header_row = idx
            break
    if header_row is None:
        raise ValueError("未找到数据表头行（Item Nbr）")
    return range_dates, header_row

def load_order_data(filepath, header_row):
    df = pd.read_excel(filepath, header=header_row, dtype=str)
    range_cols = {}
    for col in df.columns:
        match = re.match(r"Range\s*(\d+)\s*Total Units Sold", col, re.IGNORECASE)
        if match:
            rn = int(match.group(1))
            range_cols[rn] = col
    if not range_cols:
        raise ValueError("未找到任何 Range X Total Units Sold 列")
    if "Club Nbr" not in df.columns:
        raise ValueError("订单文件缺少 Club Nbr 列")
    df["Club Nbr"] = df["Club Nbr"].astype(str).str.extract(r"(\d+)")[0].astype(float).astype(int)
    keep_cols = ["Club Nbr"] + list(range_cols.values())
    df_clean = df[keep_cols].copy()
    for col in range_cols.values():
        df_clean[col] = pd.to_numeric(df_clean[col], errors="coerce").fillna(0)
    agg_dict = {col: "sum" for col in range_cols.values()}
    grouped = df_clean.groupby("Club Nbr", as_index=False).agg(agg_dict)
    rename_map = {col: f"Range{rn}_Units" for rn, col in range_cols.items()}
    grouped.rename(columns=rename_map, inplace=True)
    return grouped

def analyze_sales(order_file):
    try:
        range_dates, header_row = parse_order_metadata(order_file)
        if 1 not in range_dates:
            raise ValueError("未找到 Range 1 的时间范围，无法确定销售日期")
        sale_date = range_dates[1]["end"]
        sale_date_str = sale_date.strftime("%Y年%m月%d日")

        order_df = load_order_data(order_file, header_row)

        total_units = {}
        for col in order_df.columns:
            if col.startswith("Range") and col.endswith("_Units"):
                rn = int(col.split("Range")[1].split("_")[0])
                total_units[rn] = order_df[col].sum()

        day_sales = total_units.get(1, 0)

        if day_sales == 0:
            best_store_desc = "无销售数据"
            best_units = 0
        else:
            store_sales = order_df[["Club Nbr", "Range1_Units"]].copy()
            max_row = store_sales.loc[store_sales["Range1_Units"].idxmax()]
            best_club = max_row["Club Nbr"]
            best_units = max_row["Range1_Units"]
            store_type, province, erp = get_store_info(best_club)
            best_store_desc = f"{store_type}（{province}）{erp}"

        def growth_rate(current, previous):
            if previous == 0:
                return float('inf') if current > 0 else 0
            return (current - previous) / previous * 100

        month_growth = growth_rate(total_units.get(2, 0), total_units.get(3, 0))
        year_growth = growth_rate(total_units.get(4, 0), total_units.get(5, 0))

        # 门店明细
        store_list = []
        for _, row in order_df.iterrows():
            club = row["Club Nbr"]
            units = row.get("Range1_Units", 0)
            if units == 0:
                continue
            store_type, province, erp = get_store_info(club)
            club_name = erp
            store_list.append({
                "Club Nbr": club,
                "门店名称": club_name,
                "类型": store_type,
                "省份": province,
                "Range1销量": int(units)
            })
        store_list.sort(key=lambda x: x["Range1销量"], reverse=True)

        # 省份汇总
        province_sales = {}
        for s in store_list:
            prov = s["省份"]
            province_sales[prov] = province_sales.get(prov, 0) + s["Range1销量"]
        # 按销量从高到低排序
        sorted_provinces = sorted(province_sales.items(), key=lambda x: x[1], reverse=True)
        
        # 取前五名门店
        top5_stores = store_list[:5]

        result_lines = []
        result_lines.append("")
        result_lines.append("=" * 60)
        result_lines.append(f"订单文件：{order_file}")
        result_lines.append("门店信息：内置字典")
        result_lines.append("=" * 60)
        result_lines.append("")
        result_lines.append(f"销售日期：{sale_date_str}  总销量：{day_sales:,.0f} 瓶")
        result_lines.append("")
        if best_units > 0:
            result_lines.append(f"最高销售门店：{best_store_desc}")
            result_lines.append(f"销售瓶数：{best_units:,.0f} 瓶")
        else:
            result_lines.append("无单日销售数据，无法确定最高销售门店。")
        result_lines.append("")
        result_lines.append(f"较去年同月份销量成长率：{month_growth:+.2f}% （对比 Range2 vs Range3）")
        result_lines.append(f"较去年同年份销量成长率：{year_growth:+.2f}% （对比 Range4 vs Range5）")
        result_lines.append("")
        result_lines.append("【前五名门店销量排名】")
        result_lines.append("-" * 40)
        for i, store in enumerate(top5_stores, 1):
            result_lines.append(f"  第{i}名：{store['门店名称']} ({store['省份']}) - {store['Range1销量']:,} 瓶")
        result_lines.append("")
        result_lines.append("【省份销量排名】")
        result_lines.append("-" * 40)
        for i, (prov, sales) in enumerate(sorted_provinces, 1):
            result_lines.append(f"  第{i}名：{prov} - {sales:,} 瓶")
        result_lines.append("")
        result_lines.append("【各时间段销量与周期】")
        all_ranges = sorted(set(total_units.keys()) | set(range_dates.keys()))
        for rn in all_ranges:
            units = total_units.get(rn, 0)
            if rn in range_dates:
                start = range_dates[rn]["start"].strftime("%Y-%m-%d")
                end = range_dates[rn]["end"].strftime("%Y-%m-%d")
                period = f"{start} 至 {end}"
            else:
                period = "时间范围未知"
            result_lines.append(f"  Range{rn}：{units:,.0f} 瓶  （{period}）")
        return "\n".join(result_lines), None
    except Exception as e:
        return None, str(e)

# ==================== 导出优化 Excel ====================
def export_optimized_excel(order_file, output_path):
    try:
        _, header_row = parse_order_metadata(order_file)
        df_full = pd.read_excel(order_file, header=header_row, dtype=str)

        if "Club Nbr" not in df_full.columns:
            raise ValueError("订单文件缺少 Club Nbr 列")
        df_full["Club Nbr"] = df_full["Club Nbr"].astype(str).str.extract(r"(\d+)")[0].astype(float).astype(int)

        # ---------- 改进的列识别逻辑（支持库存、在途）----------
        col_map = {}  # { range_num: {"units": col_name, "sales": col_name, "oh": col_name, "oo": col_name} }
        # 定义可能出现的指标关键词
        metric_keywords = {
            "units": ["Total Units Sold", "total units sold"],
            "sales": ["Total Sell Dollars", "total sell dollars"],
            "oh": ["Current On-Hand Qty", "current on-hand qty", "On-Hand Qty"],
            "oo": ["Current On-Order Qty", "current on-order qty", "On-Order Qty"]
        }
        
        for col in df_full.columns:
            # 跳过 POS 相关列
            if "pos" in col.lower() or "club count" in col.lower():
                continue
            # 尝试匹配 Range X 后面的内容
            # 正则：Range 后跟数字，然后任意空白，再跟指标描述
            m = re.match(r"Range\s*(\d+)\s*(.+)", col, re.IGNORECASE)
            if m:
                rn = int(m.group(1))
                remainder = m.group(2).strip()
                # 判断 remainder 属于哪个指标
                for metric, keywords in metric_keywords.items():
                    if any(kw in remainder for kw in keywords):
                        if rn not in col_map:
                            col_map[rn] = {}
                        col_map[rn][metric] = col
                        break

        if not col_map:
            raise ValueError("未能识别任何 Range 相关列（销量、销售额、库存、在途）")

        # 构建聚合字典
        agg_dict = {}
        for rn, cols in col_map.items():
            for metric in ["units", "sales", "oh", "oo"]:
                if metric in cols:
                    agg_dict[cols[metric]] = "sum"

        # 转换数值列
        for col in agg_dict.keys():
            if col not in ["Club Nbr"]:
                df_full[col] = pd.to_numeric(df_full[col], errors="coerce").fillna(0)

        # 按门店聚合
        grouped = df_full.groupby("Club Nbr", as_index=False).agg(agg_dict)

        # 添加门店信息
        store_info = grouped["Club Nbr"].apply(lambda x: get_store_info(x))
        grouped["类型"] = store_info.apply(lambda x: x[0])
        grouped["省份"] = store_info.apply(lambda x: x[1])
        grouped["ERP"] = store_info.apply(lambda x: x[2])

        if "Club Name" in df_full.columns:
            club_names = df_full.groupby("Club Nbr")["Club Name"].first().reset_index()
            grouped = grouped.merge(club_names, on="Club Nbr", how="left")
            grouped.rename(columns={"Club Name": "门店名称"}, inplace=True)
        else:
            grouped["门店名称"] = ""

        # ---------- 省份汇总 ----------
        df_full["省份"] = df_full["Club Nbr"].apply(lambda x: get_store_info(x)[1])
        province_agg = {col: "sum" for col in agg_dict.keys()}
        prov_grouped = df_full.groupby("省份", as_index=False).agg(province_agg)

        # 重命名列为中文（统一命名）
        rename_dict = {}
        for rn, cols in col_map.items():
            if "units" in cols:
                rename_dict[cols["units"]] = f"Range{rn}_销量"
            if "sales" in cols:
                rename_dict[cols["sales"]] = f"Range{rn}_销售额"
            # 只有 Range1 有 库存 和 在途
            if rn == 1:
                if "oh" in cols:
                    rename_dict[cols["oh"]] = f"Range{rn}_库存"
                if "oo" in cols:
                    rename_dict[cols["oo"]] = f"Range{rn}_在途"
        grouped.rename(columns=rename_dict, inplace=True)
        prov_grouped.rename(columns=rename_dict, inplace=True)

        # 成长率计算函数
        def growth_rate(current, previous):
            if pd.isna(previous) or pd.isna(current):
                return 0
            if previous == 0:
                return None if current > 0 else 0
            return (current - previous) / previous * 100

        # 为门店 Dataframe 计算成长率（仅销量和销售额）
        if 2 in col_map and 3 in col_map and "units" in col_map[2] and "units" in col_map[3]:
            grouped["Range2-3_销量成长率"] = grouped.apply(
                lambda row: growth_rate(row.get(f"Range{2}_销量", 0), row.get(f"Range{3}_销量", 0)), axis=1)
        if 4 in col_map and 5 in col_map and "units" in col_map[4] and "units" in col_map[5]:
            grouped["Range4-5_销量成长率"] = grouped.apply(
                lambda row: growth_rate(row.get(f"Range{4}_销量", 0), row.get(f"Range{5}_销量", 0)), axis=1)
        if 2 in col_map and 3 in col_map and "sales" in col_map[2] and "sales" in col_map[3]:
            grouped["Range2-3_销售额成长率"] = grouped.apply(
                lambda row: growth_rate(row.get(f"Range{2}_销售额", 0), row.get(f"Range{3}_销售额", 0)), axis=1)
        if 4 in col_map and 5 in col_map and "sales" in col_map[4] and "sales" in col_map[5]:
            grouped["Range4-5_销售额成长率"] = grouped.apply(
                lambda row: growth_rate(row.get(f"Range{4}_销售额", 0), row.get(f"Range{5}_销售额", 0)), axis=1)

        # 为省份 Dataframe 计算成长率
        if 2 in col_map and 3 in col_map and "units" in col_map[2] and "units" in col_map[3]:
            prov_grouped["Range2-3_销量成长率"] = prov_grouped.apply(
                lambda row: growth_rate(row.get(f"Range{2}_销量", 0), row.get(f"Range{3}_销量", 0)), axis=1)
        if 4 in col_map and 5 in col_map and "units" in col_map[4] and "units" in col_map[5]:
            prov_grouped["Range4-5_销量成长率"] = prov_grouped.apply(
                lambda row: growth_rate(row.get(f"Range{4}_销量", 0), row.get(f"Range{5}_销量", 0)), axis=1)
        if 2 in col_map and 3 in col_map and "sales" in col_map[2] and "sales" in col_map[3]:
            prov_grouped["Range2-3_销售额成长率"] = prov_grouped.apply(
                lambda row: growth_rate(row.get(f"Range{2}_销售额", 0), row.get(f"Range{3}_销售额", 0)), axis=1)
        if 4 in col_map and 5 in col_map and "sales" in col_map[4] and "sales" in col_map[5]:
            prov_grouped["Range4-5_销售额成长率"] = prov_grouped.apply(
                lambda row: growth_rate(row.get(f"Range{4}_销售额", 0), row.get(f"Range{5}_销售额", 0)), axis=1)

        # 为省份添加最好/最差门店（基于 Range1 销量）
        if "Range1_销量" in grouped.columns:
            prov_best = {}
            prov_worst = {}
            for prov, sub_df in grouped.groupby("省份"):
                if not sub_df.empty:
                    best_row = sub_df.loc[sub_df["Range1_销量"].idxmax()]
                    worst_row = sub_df.loc[sub_df["Range1_销量"].idxmin()]
                    prov_best[prov] = f"{best_row['ERP']} ({int(best_row['Range1_销量'])}瓶)"
                    prov_worst[prov] = f"{worst_row['ERP']} ({int(worst_row['Range1_销量'])}瓶)"
            prov_grouped["销量最好门店(ERP)"] = prov_grouped["省份"].map(prov_best).fillna("")
            prov_grouped["销量最差门店(ERP)"] = prov_grouped["省份"].map(prov_worst).fillna("")

        # 调整列序（确保库存和在途列出现在对应的 Range 组中）
        base_cols = ["Club Nbr", "门店名称", "类型", "省份", "ERP"]
        # 提取所有以 Range 开头的列（不包括成长率），并按 Range 数字排序
        all_range_cols = [c for c in grouped.columns if c.startswith("Range") and "成长率" not in c]
        # 自定义排序：先按 Range数字，再按指标类型（销量、销售额、库存、在途）
        def sort_key(col):
            # 匹配 Range数字 和 后面的指标
            m = re.search(r"Range(\d+)_(.+)", col)
            if m:
                rn = int(m.group(1))
                metric = m.group(2)
                order = {"销量": 1, "销售额": 2, "库存": 3, "在途": 4}
                return (rn, order.get(metric, 99))
            return (999, col)
        all_range_cols.sort(key=sort_key)
        growth_cols = [c for c in grouped.columns if "成长率" in c]
        final_cols = base_cols + all_range_cols + growth_cols
        grouped = grouped[final_cols]
        if "Range1_销量" in grouped.columns:
            grouped = grouped.sort_values("Range1_销量", ascending=False)

        # 省份汇总列序
        prov_base_cols = ["省份"]
        prov_range_cols = [c for c in prov_grouped.columns if c.startswith("Range") and "成长率" not in c]
        prov_range_cols.sort(key=sort_key)
        prov_growth_cols = [c for c in prov_grouped.columns if "成长率" in c]
        prov_final_cols = prov_base_cols + prov_range_cols + prov_growth_cols + ["销量最好门店(ERP)", "销量最差门店(ERP)"]
        prov_grouped = prov_grouped[prov_final_cols]
        if "Range1_销量" in prov_grouped.columns:
            prov_grouped = prov_grouped.sort_values("Range1_销量", ascending=False)

        # 写入 Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            grouped.to_excel(writer, sheet_name='门店汇总', index=False, startrow=1)
            prov_grouped.to_excel(writer, sheet_name='省份汇总', index=False, startrow=1)

            # ---------- 以下为格式化函数（保持不变，但确保库存/在途列被正确格式化）----------
            def create_multi_header(ws, df):
                # 找到基础列结束位置
                base_cols_end = 0
                for i, col in enumerate(df.columns):
                    if col.startswith("Range") or "成长率" in col:
                        base_cols_end = i
                        break
                    else:
                        base_cols_end = i + 1
                # 第一行基础列留空
                for col_idx in range(1, base_cols_end + 1):
                    ws.cell(row=1, column=col_idx, value="")
                    ws.cell(row=1, column=col_idx).font = Font(bold=True)
                # 收集所有 Range 数字
                range_nums = set()
                for col in df.columns[base_cols_end:]:
                    if col.startswith("Range") and "成长率" not in col:
                        m = re.search(r"Range(\d+)", col)
                        if m:
                            range_nums.add(int(m.group(1)))
                sorted_ranges = sorted(range_nums)
                # 构建列分组
                col_groups = {}
                current_col = base_cols_end + 1
                for rn in sorted_ranges:
                    start_col = current_col
                    # 该 Range 包含的列（销量、销售额、库存、在途）
                    for col in df.columns[base_cols_end:]:
                        if col.startswith(f"Range{rn}_"):
                            current_col += 1
                    col_groups[f"Range{rn}"] = (start_col, current_col - 1)
                # 成长率分组
                growth_start_col = current_col
                for col in df.columns[base_cols_end:]:
                    if "成长率" in col:
                        current_col += 1
                if growth_start_col < current_col:
                    col_groups["成长率"] = (growth_start_col, current_col - 1)
                # 合并单元格并写入分组名
                for group_name, (start, end) in col_groups.items():
                    if start == end:
                        ws.cell(row=1, column=start, value=group_name)
                        ws.cell(row=1, column=start).font = Font(bold=True)
                        ws.cell(row=1, column=start).alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                        ws.cell(row=1, column=start, value=group_name)
                        ws.cell(row=1, column=start).font = Font(bold=True)
                        ws.cell(row=1, column=start).alignment = Alignment(horizontal='center', vertical='center')

            def format_sheet(ws, df):
                create_multi_header(ws, df)
                # 第二行（实际列名）加粗居中
                for cell in ws[2]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # 设置列宽和对齐
                for col_idx, col_name in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for cell in ws[col_letter]:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            width = 0
                            for ch in cell_str:
                                if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                                    width += 2
                                else:
                                    width += 1
                            if width > max_len:
                                max_len = width
                    adjusted_width = min(max(max_len + 2, 10), 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                    # 对齐方式：数值类右对齐，其他左对齐
                    align = 'right' if any(kw in col_name for kw in ['销量', '销售额', '库存', '在途', '成长率', 'Nbr']) else 'left'
                    for row in range(3, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.alignment = Alignment(horizontal=align, vertical='center')
                        if "成长率" in col_name and cell.value is not None:
                            if isinstance(cell.value, (int, float)):
                                cell.value = f"{cell.value:.2f}%"
                            elif cell.value is None:
                                cell.value = "∞"

            # 格式化门店汇总并添加总计行
            ws_store = writer.sheets['门店汇总']
            format_sheet(ws_store, grouped)
            # 确定门店汇总的数值列（排除基础列和成长率）
            numeric_cols_store = [c for c in grouped.columns if c not in base_cols and "成长率" not in c]
            sum_row_store = ws_store.max_row + 1
            ws_store.cell(row=sum_row_store, column=1, value="总计")
            ws_store.cell(row=sum_row_store, column=1).font = Font(bold=True)
            total_col_indices = {}
            for col_idx, col_name in enumerate(grouped.columns, start=1):
                if col_name in numeric_cols_store:
                    # 使用 Excel SUM 公式
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}3:{col_letter}{ws_store.max_row})"
                    ws_store.cell(row=sum_row_store, column=col_idx, value=formula)
                    ws_store.cell(row=sum_row_store, column=col_idx).font = Font(bold=True)
                    ws_store.cell(row=sum_row_store, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')
                    total_col_indices[col_name] = col_idx
            # 计算成长率列的总计值（也使用公式）
            for col_idx, col_name in enumerate(grouped.columns, start=1):
                if "成长率" in col_name:
                    if "Range2-3" in col_name:
                        if "销量" in col_name:
                            cur_col_idx = total_col_indices.get("Range2_销量")
                            prev_col_idx = total_col_indices.get("Range3_销量")
                        else:
                            cur_col_idx = total_col_indices.get("Range2_销售额")
                            prev_col_idx = total_col_indices.get("Range3_销售额")
                    elif "Range4-5" in col_name:
                        if "销量" in col_name:
                            cur_col_idx = total_col_indices.get("Range4_销量")
                            prev_col_idx = total_col_indices.get("Range5_销量")
                        else:
                            cur_col_idx = total_col_indices.get("Range4_销售额")
                            prev_col_idx = total_col_indices.get("Range5_销售额")
                    else:
                        continue
                    if cur_col_idx and prev_col_idx:
                        cur_letter = get_column_letter(cur_col_idx)
                        prev_letter = get_column_letter(prev_col_idx)
                        # 公式 =IF(prev=0,IF(cur>0,"∞","0.00%"),((cur-prev)/prev*100))
                        formula = f'=IF({prev_letter}{sum_row_store}=0,IF({cur_letter}{sum_row_store}>0,"∞","0.00%"),(({cur_letter}{sum_row_store}-{prev_letter}{sum_row_store})/{prev_letter}{sum_row_store}*100))'
                        ws_store.cell(row=sum_row_store, column=col_idx, value=formula)
                        ws_store.cell(row=sum_row_store, column=col_idx).font = Font(bold=True)
                        ws_store.cell(row=sum_row_store, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')

            # 格式化省份汇总并添加总计行
            ws_prov = writer.sheets['省份汇总']
            format_sheet(ws_prov, prov_grouped)
            prov_base_cols = ["省份"]
            numeric_cols_prov = [c for c in prov_grouped.columns if c not in prov_base_cols and "成长率" not in c and "销量最好门店" not in c and "销量最差门店" not in c]
            sum_row_prov = ws_prov.max_row + 1
            ws_prov.cell(row=sum_row_prov, column=1, value="总计")
            ws_prov.cell(row=sum_row_prov, column=1).font = Font(bold=True)
            total_col_indices_prov = {}
            for col_idx, col_name in enumerate(prov_grouped.columns, start=1):
                if col_name in numeric_cols_prov:
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}3:{col_letter}{ws_prov.max_row})"
                    ws_prov.cell(row=sum_row_prov, column=col_idx, value=formula)
                    ws_prov.cell(row=sum_row_prov, column=col_idx).font = Font(bold=True)
                    ws_prov.cell(row=sum_row_prov, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')
                    total_col_indices_prov[col_name] = col_idx
            for col_idx, col_name in enumerate(prov_grouped.columns, start=1):
                if "成长率" in col_name:
                    if "Range2-3" in col_name:
                        if "销量" in col_name:
                            cur_col_idx = total_col_indices_prov.get("Range2_销量")
                            prev_col_idx = total_col_indices_prov.get("Range3_销量")
                        else:
                            cur_col_idx = total_col_indices_prov.get("Range2_销售额")
                            prev_col_idx = total_col_indices_prov.get("Range3_销售额")
                    elif "Range4-5" in col_name:
                        if "销量" in col_name:
                            cur_col_idx = total_col_indices_prov.get("Range4_销量")
                            prev_col_idx = total_col_indices_prov.get("Range5_销量")
                        else:
                            cur_col_idx = total_col_indices_prov.get("Range4_销售额")
                            prev_col_idx = total_col_indices_prov.get("Range5_销售额")
                    else:
                        continue
                    if cur_col_idx and prev_col_idx:
                        cur_letter = get_column_letter(cur_col_idx)
                        prev_letter = get_column_letter(prev_col_idx)
                        formula = f'=IF({prev_letter}{sum_row_prov}=0,IF({cur_letter}{sum_row_prov}>0,"∞","0.00%"),(({cur_letter}{sum_row_prov}-{prev_letter}{sum_row_prov})/{prev_letter}{sum_row_prov}*100))'
                        ws_prov.cell(row=sum_row_prov, column=col_idx, value=formula)
                        ws_prov.cell(row=sum_row_prov, column=col_idx).font = Font(bold=True)
                        ws_prov.cell(row=sum_row_prov, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')

            # 时间范围说明 Sheet
            range_dates, _ = parse_order_metadata(order_file)
            if range_dates:
                meta_df = pd.DataFrame([
                    {"时间段": rn, "开始日期": info["start"].strftime("%Y-%m-%d"), "结束日期": info["end"].strftime("%Y-%m-%d")}
                    for rn, info in range_dates.items()
                ])
                meta_df.to_excel(writer, sheet_name="时间范围说明", index=False)
                ws_meta = writer.sheets["时间范围说明"]
                for cell in ws_meta[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for col_idx, col_name in enumerate(meta_df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for cell in ws_meta[col_letter]:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            width = 0
                            for ch in cell_str:
                                if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                                    width += 2
                                else:
                                    width += 1
                            if width > max_len:
                                max_len = width
                    adjusted_width = min(max(max_len + 2, 10), 50)
                    ws_meta.column_dimensions[col_letter].width = adjusted_width
                    for row in range(2, ws_meta.max_row + 1):
                        ws_meta.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

        return True, None
    except Exception as e:
        return False, str(e)
# ==================== GUI ====================
class SalesAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("hefengfan山姆销售分析")
        self.root.geometry("500x600")
        self.root.option_add('*Font', ('Consolas', 10))

        self.order_path = tk.StringVar()
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self.root, text="订单文件（必选）:", font=("微软雅黑", 10)).pack(pady=(10,0), anchor="w", padx=10)
        frame_order = tk.Frame(self.root)
        frame_order.pack(fill="x", padx=10, pady=5)
        tk.Entry(frame_order, textvariable=self.order_path, width=40, font=("微软雅黑", 10)).pack(side="left", fill="x", expand=True)
        tk.Button(frame_order, text="浏览...", command=self.browse_order).pack(side="right", padx=5)

        self.analyze_btn = tk.Button(self.root, text="开始分析", command=self.analyze, width=20, font=("微软雅黑", 10))
        self.analyze_btn.pack(pady=10)

        self.export_btn = tk.Button(self.root, text="导出优化 Excel", command=self.export_excel, width=20, font=("微软雅黑", 10))
        self.export_btn.pack(pady=5)

        self.result_text = scrolledtext.ScrolledText(self.root, wrap=tk.NONE, width=70, height=30, font=("Consolas", 10))
        self.result_text.pack(padx=10, pady=5, fill="both", expand=True)

    def browse_order(self):
        fname = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
        if fname:
            self.order_path.set(fname)

    def analyze(self):
        order_file = self.order_path.get().strip()
        if not order_file:
            messagebox.showerror("错误", "请先选择订单文件")
            return

        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, "正在分析，请稍候...\n")
        self.root.update()

        result, error = analyze_sales(order_file)
        if error:
            messagebox.showerror("分析失败", error)
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, f"发生错误：\n{error}")
        else:
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, result)

    def export_excel(self):
        order_file = self.order_path.get().strip()
        if not order_file:
            messagebox.showerror("错误", "请先选择订单文件")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="保存优化Excel报表"
        )
        if not output_path:
            return

        self.result_text.insert(tk.END, "\n正在导出优化Excel，请稍候...\n")
        self.root.update()

        success, error = export_optimized_excel(order_file, output_path)
        if success:
            messagebox.showinfo("导出成功", f"报表已保存至：\n{output_path}")
            self.result_text.insert(tk.END, f"\n✅ 优化Excel导出成功：{output_path}\n")
        else:
            messagebox.showerror("导出失败", error)
            self.result_text.insert(tk.END, f"\n❌ 导出失败：{error}\n")

if __name__ == "__main__":
    root = tk.Tk()
    app = SalesAnalyzerGUI(root)
    root.mainloop()
